import streamlit as st
import json
import pandas as pd
import numpy as np
import plotly.graph_objects as go
import plotly.express as px
from pathlib import Path
import math
import io
import base64
import requests
from datetime import datetime
from urllib.parse import quote

st.set_page_config(page_title="Puzzle Creator", layout="wide", page_icon="🧩")

# ══════════════════════════════════════════════════════
# 상수
# ══════════════════════════════════════════════════════
BASE       = Path(__file__).parent
LEVELS_DIR = BASE / "data" / "levels"
INTG_CSV   = BASE / "data" / "integrated_difficulty.csv"
TBLSTAGE   = BASE / "data" / "tblStage_500.xlsx"
ARCHIVE_DIR= BASE / "data" / "archives"

COLOR_MAP  = {0:'Blue',1:'Yellow',2:'Red',3:'Green',4:'Orange',5:'Purple',6:'White',7:'Black'}
HEX_COLORS = {
    'Normal':'#D0D0D0','Blank':'#1a1a2e','Stack':'#4A90D9',
    'Lock':'#5C5C5C','Plank':'#8B5E3C','Ice':'#A8D8EA',
    'StackLock':'#6A4C93','Grass':'#52C41A','Ads':'#FA8C16',
    'CameraPicture':'#EB2F96',
}
CHIP_HEX   = {0:'#1890FF',1:'#FADB14',2:'#F5222D',3:'#52C41A',
              4:'#FA8C16',5:'#722ED1',6:'#FAFAFA',7:'#141414'}
TILETYPE   = {0:'Normal',1:'Blank',2:'Stack',3:'Lock',4:'Plank',
              5:'Ice',6:'StackLock',7:'Grass',8:'Ads',9:'CameraPicture'}
TILE_REV   = {v:k for k,v in TILETYPE.items()}
# ── H1 지표 컬럼 매핑 (market_df 컬럼명 대응)
MK_COL_MAP = {
    "H1_1":"H1-1","H1_2":"H1-2","H1_3":"H1-3 ",
    "H1_4":"H1-4","H1_5":"H1-5","H1_6":"H1-6 ",
    "H1_7":"H1-7 ","H1_8":"H1-8 ","H1_9":"H1-9 ",
    "H1_12":"H1-12 ","H1_13":"H1-13 ","H1_14":"H1-14",
}

# ── H1 지표 난이도 방향 (True = 역수, 높을수록 쉬움)
W_DIR = {
    "H1_1":True,"H1_2":True,"H1_3":True,"H1_4":True,"H1_5":False,
    "H1_6":False,"H1_7":False,"H1_8":False,"H1_9":False,"H1_10":False,
    "H1_11":False,"H1_12":False,"H1_13":True,"H1_14":True,"H1_15":True,
}

GITHUB_REPO  = "Ahyeong0202/My-Little-Bookstore_Puzzle-Creator"
ARCHIVE_PATH = "data/archives"
MARKET_CSV   = BASE / "data" / "market" / "market_lv1_100.csv"
ASSETS_IMG   = BASE / "assets" / "images"
ASSETS_VID   = BASE / "assets" / "videos"
GITHUB_RAW   = f"https://raw.githubusercontent.com/{GITHUB_REPO}/main"


# ══════════════════════════════════════════════════════
# 번역 딕셔너리 (한국어 / English)
# 사용법: TR["key"] → IS_EN 기준으로 자동 선택
# ══════════════════════════════════════════════════════
_TR = {
    # ── 사이드바
    "page":             ("페이지",                      "Page"),
    "file_upload":      ("📂 파일 업로드",               "📂 File Upload"),
    "download":         ("📥 다운로드",                  "📥 Download"),
    "csv_intg":         ("CSV — 통합 난이도",            "CSV — Integrated Difficulty"),
    "json_intg":        ("JSON — 통합 난이도",           "JSON — Integrated Difficulty"),
    "upload_hint":      ("파일 업로드 후 다운로드 가능",   "Available after file upload"),
    "github_settings":  ("🔑 GitHub 설정",               "🔑 GitHub Settings"),
    "market_upload":    ("시장 데이터 CSV (새로 업로드)", "Market Data CSV (Upload New)"),
    "tbl_auto":         ("GitHub 자동 로드",             "GitHub"),
    "tbl_missing":      ("⚠ tblStage 없음",             "⚠ No tblStage"),
    # ── 홈 탭
    "hero_sub":         ("헥사소트 퍼즐 레벨 난이도 설계 자동화 시스템", "Hexasort Puzzle Level Difficulty Design System"),
    "metric_levels":    ("퍼즐 레벨",                   "Puzzle Levels"),
    "metric_market":    ("시장 데이터",                  "Market Data"),
    "metric_metrics":   ("난이도 지표",                  "Difficulty Metrics"),
    "metric_launch":    ("출시 목표",                    "Launch Target"),
    "sec_video":        ("🎮 게임 플레이 영상",           "🎮 Gameplay Video"),
    "sec_intro":        ("📖 게임 소개",                 "📖 Game Introduction"),
    # ── 매뉴얼 탭
    "tab_manual":       ("📖 1. 매뉴얼",                "📖 1. Manual"),
    "tab_analysis":     ("📊 2. 난이도 분석",            "📊 2. Difficulty Analysis"),
    "tab_viewer":       ("🗺️ 3. 판 모양 뷰어",          "🗺️ 3. Board Viewer"),
    "tab_generator":    ("🎲 4. JSON 생성기",            "🎲 4. JSON Generator"),
    "tab_settings":     ("🔧 5. 설정",                  "🔧 5. Settings"),
    "tab_archive":      ("🗄️ 6. 아카이브",              "🗄️ 6. Archive"),
    "tab_guide":        ("## 탭 안내",                  "## Tab Guide"),
    "file_structure":   ("## 파일 구조",                "## File Structure"),
    "update_json":      ("## JSON 업데이트 방법",        "## How to Update JSONs"),
    "glossary":         ("## 용어 사전",                "## Glossary"),
    "tiletype_exp":     ("TileType — 셀 종류",          "TileType — Cell Types"),
    "h1_exp":           ("H1 지표 — 판 모양 난이도 수치","H1 Indicators — Board Difficulty Metrics"),
    "tbl_exp":          ("tblStage 파라미터",            "tblStage Parameters"),
    "chip_exp":         ("칩 색상 코드",                "Chip Color Codes"),
    "grade_exp":        ("난이도 등급",                  "Difficulty Grades"),
    "formula_exp":      ("난이도 곡선 공식",             "Difficulty Curve Formula"),
    # ── 난이도 분석 탭
    "analysis_title":   ("📊 난이도 분석",              "📊 Difficulty Analysis"),
    "analysis_cap":     ("시장 데이터 기준선 vs 우리 게임 통합 난이도 비교 + 가중치 조정",
                         "Market data baseline vs our game integrated difficulty + weight adjustment"),
    "market_exp":       ("📋 시장 데이터 원본 (Lv 1~100) — 클릭하여 펼치기",
                         "📋 Market Data Raw (Lv 1~100) — Click to expand"),
    "h1_trend":         ("**H1 지표별 원시값 추이**",   "**H1 Indicator Raw Trends**"),
    "show_indicators":  ("표시할 지표",                 "Show Indicators"),
    "weight_adj":       ("⚖️ 판:게임진행 가중치 조정",  "⚖️ Board:Gameplay Weight Adjustment"),
    "board_weight":     ("판 모양 가중치 (%)",           "Board weight (%)"),
    "select_curves":    ("**표시할 곡선 선택**",         "**Select Curves**"),
    "our_intg":         ("우리 통합",                   "Our Integrated"),
    "market_board":     ("시장 board",                  "Market Board"),
    "difficulty_curve": ("📈 난이도 곡선 비교",          "📈 Difficulty Curve Comparison"),
    "zone_analysis":    ("📉 우리 난이도 구간 분석",     "📉 Our Difficulty Zone Analysis"),
    "zone_size":        ("Zone Size" ,                  "Zone Size"),
    "avg_intg":         ("평균 통합 난이도",             "Avg Integrated Difficulty"),
    # ── 판 모양 뷰어
    "viewer_title":     ("🗺️ 판 모양 뷰어 / JSON 생성기","🗺️ Board Viewer / JSON Generator"),
    "lv_viewer":        ("🔍 레벨 뷰어 & 편집",         "🔍 Level Viewer & Edit"),
    "new_board":        ("✏️ 새 판 만들기",              "✏️ New Board"),
    "source":           ("소스",                        "Source"),
    "lv_number":        ("레벨 번호",                   "Level Number"),
    "difficulty":       ("난이도 선택",                  "Difficulty"),
    "json_upload":      ("JSON 업로드",                  "JSON Upload"),
    "edit_mode":        ("편집 모드",                    "Edit Mode"),
    "show_coord":       ("좌표 표시",                    "Show Coordinates"),
    "show_chip":        ("칩 색상 표시",                 "Show Chip Colors"),
    "hex_size":         ("헥사 크기",                    "Hex Size"),
    "h1_csv":           ("H1 CSV",                      "H1 CSV"),
    "tile_legend":      ("타일 범례",                    "Tile Legend"),
    "apply_cell":       ("✅ 셀 적용",                   "✅ Apply Cell"),
    "reset_grid":       ("🔄 그리드 초기화",             "🔄 Reset Grid"),
    "row_sel":          ("행 (Y, 위→아래)",              "Row (Y)"),
    "col_sel":          ("열 (X, 좌→우)",               "Column (X)"),
    # ── JSON 생성기 탭
    "gen_title":        ("🎲 JSON 생성기",               "🎲 JSON Generator"),
    "gen_cap":          ("난이도 곡선 기반으로 레벨 범위를 선택해 JSON 파일을 생성하고 다운로드합니다.",
                         "Select level range to generate and download JSON files based on difficulty curve."),
    "gen_curve":        ("📈 난이도 곡선",               "📈 Difficulty Curve"),
    "lv_range":         ("생성할 레벨 범위",             "Level Range"),
    "avg_diff":         ("평균 난이도",                  "Avg Difficulty"),
    # ── 설정 탭
    "settings_title":   ("🔧 설정",                     "🔧 Settings"),
    "h1_weight_tab":    ("⚖️ H1 가중치 설정",           "⚖️ H1 Weight Settings"),
    "tbl_weight_tab":   ("🎰 tblStage 가중치 설정",     "🎰 tblStage Weight Settings"),
    "stack_edit_tab":   ("📋 스택 파라미터 수정",        "📋 Stack Parameter Editor"),
    "h1_weight_cap":    ("각 H1 지표의 기여 비율을 설정합니다. 합계가 100%가 되도록 자동 정규화됩니다.",
                         "Set contribution ratio for each H1 indicator. Auto-normalized to 100%."),
    "tbl_weight_cap":   ("tblStage 파라미터별 게임 진행 난이도 기여 가중치를 설정합니다.",
                         "Set gameplay difficulty contribution weights per tblStage parameter."),
    "stack_cap":        ("셀을 클릭해 직접 수정할 수 있습니다. 수정 후 Excel로 다운로드하세요.",
                         "Click cells to edit directly. Download as Excel after modification."),
    # ── 아카이브 탭
    "archive_title":    ("🗄️ 아카이브",                "🗄️ Archive"),
    "archive_cap":      ("설정값을 버전으로 저장하고 GitHub에 자동 커밋합니다.",
                         "Save settings as versions and auto-commit to GitHub."),
    "save_settings":    ("💾 현재 설정 저장",           "💾 Save Current Settings"),
    "load_version":     ("📂 버전 불러오기",             "📂 Load Version"),
    "session_history":  ("📋 이번 세션 저장 기록",       "📋 Session Save History"),
    "version_compare":  ("📊 버전 비교 (난이도 곡선)",   "📊 Version Comparison (Difficulty Curve)"),
    "version_name":     ("버전명",                       "Version Name"),
    "memo":             ("메모",                         "Memo"),
    "save_local":       ("🖫 로컬에 저장 (JSON)",        "🖫 Save (JSON)"),
    "commit_github":    ("☁️ GitHub에 커밋",            "☁️ Commit to GitHub"),
    "select_version":   ("버전 선택",                    "Select Version"),
    "load_btn":         ("불러오기",                     "Load"),
}

def TR(key: str) -> str:
    """번역 헬퍼: _TR 딕셔너리에서 key에 해당하는 한/영 텍스트 반환.
    IS_EN=True면 영어, False면 한국어 반환.
    key가 없으면 key 자체를 반환.
    """
    pair = _TR.get(key, (key, key))
    return pair[1] if IS_EN else pair[0]

# ══════════════════════════════════════════════════════
# CSS
# ══════════════════════════════════════════════════════
# ── CSS는 session_state 초기화 이후에 주입 (아래 테마 적용 블록에서 처리)

# ══════════════════════════════════════════════════════
# session_state 초기화
# ══════════════════════════════════════════════════════
for key, default in {
    "intg_df":      None,   # integrated_difficulty.csv
    "tbl_df":       None,   # tblStage_500.xlsx
    "market_df":    None,   # 시장 데이터 CSV
    "level_data":   None,   # 현재 레벨 JSON
    "grid_tiles":   None,   # 3번탭 편집 중 그리드
    "grid_x":       7,
    "grid_y":       7,
    "w_board":      50,     # 판 모양 가중치 %
    "w_gameplay":   50,     # 게임 진행 가중치 %
    "github_token": "",
    "archives":     [],
    "lang":         "한국어",
}.items():
    if key not in st.session_state:
        st.session_state[key] = default

# ── 앱 시작 시 GitHub 파일 자동 로드 (업로드 없이 사용 가능)
if st.session_state.tbl_df is None and TBLSTAGE.exists():
    try:
        df_auto = pd.read_excel(TBLSTAGE, sheet_name="Stage", header=0)
        mask = (df_auto["LevelName"].str.startswith("N_", na=False) |
                df_auto["LevelName"].str.startswith("N ", na=False))
        st.session_state.tbl_df = df_auto[mask].reset_index(drop=True)
    except Exception:
        pass

if st.session_state.intg_df is None and INTG_CSV.exists():
    try:
        st.session_state.intg_df = pd.read_csv(INTG_CSV)
    except Exception:
        pass

# ── 브라운 라이트 테마 고정
T = {
    "bg":        "#FBF5EE",
    "bg2":       "#F0E6D8",
    "bg3":       "#E8D5C0",
    "border":    "#C4956A",
    "text":      "#2C1810",
    "text2":     "#7A5C45",
    "accent":    "#6B3A2A",
    "grid_bg":   "#F0E6D8",
    "plot_bg":   "#FBF5EE",
    "grid_line": "#E8D5C0",
    "brown":     "#6B3A2A",
    "brown_mid": "#8B5A3A",
    "brown_lt":  "#D4956A",
    "cream":     "#FBF5EE",
    "white":     "#FFFFFF",
}

# ── 브라운 라이트 CSS
st.markdown("""
<style>
@import url('https://fonts.googleapis.com/css2?family=Noto+Sans+KR:wght@400;500;700&display=swap');

/* ── 전체 배경 */
.stApp { background-color: #FBF5EE !important; color: #2C1810 !important; font-family: 'Noto Sans KR', sans-serif; }
[data-testid="stAppViewContainer"] { background-color: #FBF5EE !important; }
[data-testid="stHeader"] { background-color: #FBF5EE !important; border-bottom: 1px solid #E8D5C0; }
[data-testid="stMainBlockContainer"] { background-color: #FBF5EE !important; }

/* ── 사이드바 */
[data-testid="stSidebar"] { background: #F0E6D8 !important; border-right: 2px solid #C4956A; }
[data-testid="stSidebar"] * { color: #2C1810 !important; }
[data-testid="stSidebar"] .stRadio label { color: #2C1810 !important; font-weight: 500; }

/* ── 라디오 선택 */
[data-testid="stSidebar"] .stRadio [data-testid="stWidgetLabel"] { color: #2C1810 !important; }

/* ── 입력 위젯 */
[data-testid="stTextInput"] input,
[data-testid="stNumberInput"] input,
[data-testid="stSelectbox"] div,
[data-testid="stTextArea"] textarea {
    background-color: #FFFFFF !important;
    color: #2C1810 !important;
    border-color: #C4956A !important;
    border-radius: 6px !important;
}

/* ── 카드 */
.metric-card {
    background: #FFFFFF; border: 1px solid #E8D5C0;
    border-radius: 10px; padding: 16px; text-align: center;
    box-shadow: 0 2px 8px rgba(107,58,42,0.08);
}
.metric-val { font-size: 28px; font-weight: 700; color: #6B3A2A; }
.metric-lbl { font-size: 12px; color: #7A5C45; margin-top: 4px; }

/* ── 뱃지 */
.file-badge {
    display: inline-block; background: #6B3A2A;
    color: white; border-radius: 4px;
    font-size: 11px; padding: 2px 7px; margin: 2px 0;
}
.file-badge-warn {
    display: inline-block; background: #D4956A;
    color: white; border-radius: 4px;
    font-size: 11px; padding: 2px 7px; margin: 2px 0;
}

/* ── 사이드바 라벨 */
.sidebar-label {
    font-size: 11px; font-weight: 700; letter-spacing: 0.08em;
    color: #7A5C45; text-transform: uppercase; margin-bottom: 8px;
}

/* ── 탭 */
div[data-testid="stTabs"] button {
    font-size: 14px; color: #7A5C45 !important;
    border-radius: 6px 6px 0 0;
}
div[data-testid="stTabs"] button[aria-selected="true"] {
    color: #6B3A2A !important; font-weight: 700;
    border-bottom: 2px solid #6B3A2A !important;
}

/* ── 데이터프레임 */
[data-testid="stDataFrame"] { background: #FFFFFF; border: 1px solid #E8D5C0; border-radius: 8px; }

/* ── expander */
[data-testid="stExpander"] {
    background: #FFFFFF !important; border: 1px solid #E8D5C0 !important;
    border-radius: 8px;
}
[data-testid="stExpander"] summary {
    color: #2C1810 !important;
    background: #FFFFFF !important;
}
[data-testid="stExpander"] summary:hover {
    background: #F0E6D8 !important;
}
[data-testid="stExpander"] > div {
    background: #FFFFFF !important;
}

/* ── 구분선 */
hr { border-color: #E8D5C0; }

/* ── 파일 업로더 */
[data-testid="stFileUploader"] { background: #FFFFFF; border-radius: 8px; border: 1px solid #E8D5C0; }
[data-testid="stFileUploader"] * { color: #2C1810 !important; }
[data-testid="stFileUploaderDropzone"] {
    background: #F0E6D8 !important;
    border: 1.5px dashed #C4956A !important;
    border-radius: 6px;
}

/* ── 라디오/체크박스 */
[data-testid="stRadio"] label,
[data-testid="stCheckbox"] label,
[data-testid="stToggle"] label {
    color: #2C1810 !important;
}

/* ── 슬라이더 */
[data-testid="stSlider"] label { color: #2C1810 !important; }
[data-testid="stSlider"] [data-testid="stTickBarMin"],
[data-testid="stSlider"] [data-testid="stTickBarMax"] { color: #7A5C45 !important; }

/* ── 버튼 */
.stButton > button {
    background-color: #6B3A2A !important; color: #FFFFFF !important;
    border: none !important; border-radius: 8px !important; font-weight: 600;
    transition: background 0.2s;
}
.stButton > button:hover { background-color: #8B5A3A !important; }
.stButton > button * { color: #FFFFFF !important; }

/* ── 다운로드 버튼 */
[data-testid="stDownloadButton"] button {
    background-color: #6B3A2A !important; color: #FFFFFF !important;
    border: none !important; border-radius: 8px !important; font-weight: 600;
}
[data-testid="stDownloadButton"] button:hover { background-color: #8B5A3A !important; }
[data-testid="stDownloadButton"] button * { color: #FFFFFF !important; }

/* ── primary 버튼 */
[data-testid="stBaseButton-primary"] {
    background-color: #6B3A2A !important; color: #FFFFFF !important;
}

/* ── 라디오 버튼 선택된 항목 */
[data-testid="stSidebar"] .stRadio input:checked + label { color: #6B3A2A !important; font-weight: 700; }

/* ── 토글 */
[data-testid="stToggle"] { accent-color: #6B3A2A; }

/* ── 체크박스 */
[data-testid="stCheckbox"] input { accent-color: #6B3A2A; }

/* ── 슬라이더 thumb */
[data-testid="stSlider"] [role="slider"] { background-color: #6B3A2A !important; }

/* ── selectbox 선택 영역 */
[data-testid="stSelectbox"] > div > div {
    background-color: #FFFFFF !important; color: #2C1810 !important;
    border: 1px solid #C4956A !important;
}

/* ── multiselect 태그 */
[data-testid="stMultiSelect"] span[data-baseweb="tag"] {
    background-color: #6B3A2A !important; color: #FFFFFF !important;
}
[data-testid="stMultiSelect"] span[data-baseweb="tag"] span { color: #FFFFFF !important; }

/* ── number input 버튼 */
[data-testid="stNumberInput"] button {
    background-color: #F0E6D8 !important; color: #2C1810 !important;
    border-color: #C4956A !important;
}

/* ── file uploader 버튼 */
[data-testid="stFileUploaderDropzone"] button {
    background-color: #6B3A2A !important; color: #FFFFFF !important;
    border: none !important;
}
[data-testid="stFileUploaderDropzone"] button * { color: #FFFFFF !important; }
[data-testid="stFileUploaderDropzone"] button p { color: #FFFFFF !important; }
[data-testid="stFileUploaderDropzone"] button span { color: #FFFFFF !important; }
[data-testid="stFileUploader"] button {
    background-color: #6B3A2A !important; color: #FFFFFF !important;
    border: none !important;
}
[data-testid="stFileUploader"] button * { color: #FFFFFF !important; }
[data-testid="stFileUploader"] [data-testid="stBaseButton-secondary"] {
    background-color: #6B3A2A !important; color: #FFFFFF !important;
}
[data-testid="stFileUploader"] [data-testid="stBaseButton-secondary"] p,
[data-testid="stFileUploader"] [data-testid="stBaseButton-secondary"] span {
    color: #FFFFFF !important;
}

/* ── progress */
[data-testid="stProgressBar"] div { background-color: #6B3A2A !important; }

/* ── st.status */
[data-testid="stStatusWidget"] { background: #FFFFFF; border: 1px solid #E8D5C0; border-radius: 8px; }

/* ── 선택된 라디오 원 */
.stRadio [data-baseweb="radio"] div { border-color: #6B3A2A !important; }
.stRadio [data-baseweb="radio"] [data-checked="true"] div { background-color: #6B3A2A !important; }

/* ── 탭 선택 언더라인 */
[data-baseweb="tab-highlight"] { background-color: #6B3A2A !important; }

/* ── 코드 블록 */
[data-testid="stCode"] { background: #F0E6D8 !important; color: #2C1810 !important; }
[data-testid="stCode"] * { color: #2C1810 !important; }
[data-testid="stCode"] pre { background: #F0E6D8 !important; color: #2C1810 !important; }
[data-testid="stJson"] { background: #F0E6D8 !important; color: #2C1810 !important; }
[data-testid="stJson"] * { color: #2C1810 !important; background: #F0E6D8 !important; }
.stJson { background: #F0E6D8 !important; }
[data-testid="stCode"] { background: #F0E6D8 !important; }
.highlight { background: #F0E6D8 !important; }
.highlight * { background: #F0E6D8 !important; color: #2C1810 !important; }
[data-testid="stCode"] code { background: #F0E6D8 !important; color: #2C1810 !important; }
[data-testid="stCodeBlock"] { background: #F0E6D8 !important; }
[data-testid="stCodeBlock"] * { color: #2C1810 !important; background: #F0E6D8 !important; }
pre[class*="language-"] { background: #F0E6D8 !important; color: #2C1810 !important; }
code[class*="language-"] { background: #F0E6D8 !important; color: #2C1810 !important; }

/* ── alert/warning/info 텍스트 */
[data-testid="stAlert"] { background: #FBF5EE !important; border-radius: 8px; }
[data-testid="stAlert"] p, [data-testid="stAlert"] * { color: #2C1810 !important; }

/* ── ... 툴바 메뉴 (Rerun/Clear cache 등) */
[data-testid="stMainMenuPopover"],
[data-testid="stMainMenuPopover"] *,
[data-testid="stMainMenuPopover"] li,
[data-testid="stMainMenuPopover"] span,
[data-testid="stMainMenuPopover"] p,
[data-baseweb="popover"] [role="menuitem"],
[data-baseweb="popover"] [role="menuitem"] * { color: #FFFFFF !important; }

/* ── 전역: 옅은 회색 글씨 → 다크 브라운으로 강제 */
p, span, div, label, li, td, th, small {
    color: #2C1810;
}
/* plotly 범례/축 텍스트는 제외 (SVG) */
svg text { fill: #2C1810; }

/* ── 파일 업로더 Browse 버튼 전체 */
[data-testid="stFileUploaderDropzone"] button,
[data-testid="stFileUploaderDropzone"] button *,
[data-testid="stFileUploaderDropzone"] button p,
[data-testid="stFileUploaderDropzone"] button span,
[data-testid="stFileUploaderDropzoneInstructions"] button,
section[data-testid="stFileUploaderDropzone"] > div > button {
    background-color: #6B3A2A !important;
    color: #FFFFFF !important;
}

/* ── 파일 업로더 드롭존 텍스트 */
[data-testid="stFileUploaderDropzoneInstructions"] *,
[data-testid="stFileUploaderDropzoneInstructions"] span,
[data-testid="stFileUploaderDropzoneInstructions"] p {
    color: #2C1810 !important;
}

/* ── 사이드바 Browse files 버튼 */
[data-testid="stSidebar"] [data-testid="stFileUploaderDropzone"] button {
    background-color: #6B3A2A !important;
    color: #FFFFFF !important;
    border: none !important;
}
[data-testid="stSidebar"] [data-testid="stFileUploaderDropzone"] button * {
    color: #FFFFFF !important;
}
[data-testid="stSidebar"] [data-testid="stFileUploaderDropzone"] button p {
    color: #FFFFFF !important;
}

/* ── Plotly 차트 범례 텍스트 */
.js-plotly-plot .plotly .legend text { fill: #2C1810 !important; }
.js-plotly-plot .plotly .xtick text,
.js-plotly-plot .plotly .ytick text { fill: #2C1810 !important; }
.js-plotly-plot .plotly .g-xtitle text,
.js-plotly-plot .plotly .g-ytitle text { fill: #2C1810 !important; }

/* ── 탭 텍스트 */
[data-testid="stTabs"] button p,
[data-testid="stTabs"] button span { color: #2C1810 !important; }
[data-testid="stTabs"] button[aria-selected="true"] p,
[data-testid="stTabs"] button[aria-selected="true"] span {
    color: #6B3A2A !important; font-weight: 700;
}

/* ── expander 제목 */
[data-testid="stExpander"] summary p,
[data-testid="stExpander"] summary span { color: #2C1810 !important; }

/* ── select_slider / slider 값 표시 */
[data-testid="stSlider"] p,
[data-testid="stSlider"] span { color: #2C1810 !important; }

/* ── metric 델타 */
[data-testid="stMetricDelta"] { color: #6B3A2A !important; }

/* ── 아카이브 텍스트 */
[data-testid="stDataEditor"] td,
[data-testid="stDataEditor"] th { color: #2C1810 !important; background: #FFFFFF; }

/* ── caption 강화 */
.stCaption, [data-testid="stCaptionContainer"] * { color: #7A5C45 !important; }

/* ── data_editor */
[data-testid="stDataEditor"] * { color: #2C1810 !important; }
[data-testid="stDataFrameGlideDataEditor"] * { color: #2C1810 !important; }

/* ── 캡션 */
[data-testid="stCaptionContainer"] p, .stCaption { color: #7A5C45 !important; }

/* ── selectbox 드롭다운 옵션 리스트 */
[data-baseweb="popover"] ul li,
[data-baseweb="popover"] ul li span,
[data-baseweb="menu"] ul li,
[data-baseweb="menu"] ul li span,
[data-baseweb="select"] [role="option"],
[data-baseweb="select"] [role="option"] span {
    background-color: #FFFFFF !important;
    color: #2C1810 !important;
}
[data-baseweb="popover"] ul li:hover,
[data-baseweb="menu"] ul li:hover {
    background-color: #F0E6D8 !important;
}

/* ── 마크다운 */
.stMarkdown p, .stMarkdown li { color: #2C1810; }
.stMarkdown h1, .stMarkdown h2, .stMarkdown h3 { color: #6B3A2A; }
.stMarkdown code { background: #F0E6D8; color: #6B3A2A; border-radius: 4px; padding: 2px 6px; }
.stMarkdown pre { background: #F0E6D8 !important; border: 1px solid #E8D5C0; }

/* ── selectbox 라벨 */
[data-testid="stSelectbox"] label,
[data-testid="stNumberInput"] label,
[data-testid="stTextInput"] label,
[data-testid="stTextArea"] label { color: #2C1810 !important; }

/* ── progress bar */
[data-testid="stProgressBar"] > div { background-color: #6B3A2A !important; }

/* ── metric */
[data-testid="stMetric"] { background: #FFFFFF; border-radius: 8px; padding: 12px; border: 1px solid #E8D5C0; }
[data-testid="stMetricValue"] { color: #6B3A2A !important; }
[data-testid="stMetricLabel"] { color: #7A5C45 !important; }

/* ── 페이지 전환 애니메이션 */
@keyframes fadeInUp {
    from { opacity: 0; transform: translateY(24px); }
    to   { opacity: 1; transform: translateY(0); }
}
@keyframes fadeIn {
    from { opacity: 0; }
    to   { opacity: 1; }
}
@keyframes slideInLeft {
    from { opacity: 0; transform: translateX(-30px); }
    to   { opacity: 1; transform: translateX(0); }
}
.page-anim { animation: fadeInUp 0.5s ease forwards; }
.fade-in   { animation: fadeIn 0.6s ease forwards; }
.slide-in  { animation: slideInLeft 0.5s ease forwards; }

/* ── 이미지 카드 */
.img-card {
    background: #FFFFFF; border-radius: 12px;
    box-shadow: 0 4px 16px rgba(107,58,42,0.10);
    padding: 8px; margin-bottom: 12px;
    opacity: 0;
    animation: fadeInUp 0.6s ease forwards;
}

/* ── 홈 히어로 */
.hero-title {
    font-size: 2.4em; font-weight: 800; color: #6B3A2A;
    text-align: center; margin-bottom: 0.2em;
}
.hero-sub {
    font-size: 1.1em; color: #7A5C45;
    text-align: center; margin-bottom: 1.5em;
}

/* ── 섹션 헤더 */
.section-header {
    font-size: 1.4em; font-weight: 700; color: #6B3A2A;
    border-left: 4px solid #D4956A; padding-left: 12px;
    margin: 24px 0 16px 0;
}

/* ── warning/info */
[data-testid="stAlert"] { border-radius: 8px; }
</style>
""", unsafe_allow_html=True)

# ══════════════════════════════════════════════════════
# 헬퍼 함수
# ══════════════════════════════════════════════════════
def df_to_csv_bytes(df):
    """DataFrame을 UTF-8 CSV 바이트로 변환 (다운로드용)."""
    return df.to_csv(index=False).encode("utf-8-sig")

def df_to_json_bytes(df):
    """DataFrame을 JSON 바이트로 변환 (다운로드용)."""
    return json.dumps(df.to_dict(orient='records'), ensure_ascii=False, indent=2).encode("utf-8")

def hex_to_pixel(row, col, size=40):
    # flat-top 헥사, 열(col) 기준 배치
    # 가로: 열마다 size*1.5 간격
    # 세로: 행마다 size*sqrt(3) 간격, 홀수 열은 size*sqrt(3)/2 아래로 offset
    x = size * 1.5 * col
    y = -(size * math.sqrt(3) * row + (size * math.sqrt(3) / 2) * (col % 2))
    return x, y

def make_hex_path(cx, cy, size=38):
    # flat-top: 각도 0°부터 (선이 위아래, 꼭짓점 좌우)
    pts = [(cx + size * math.cos(math.pi / 180 * (60 * i)),
            cy + size * math.sin(math.pi / 180 * (60 * i))) for i in range(6)]
    pts.append(pts[0])
    return [p[0] for p in pts], [p[1] for p in pts]

# ── H1 분석 (level_analyzer_v2 인라인)
NEIGHBORS_EVEN = [(-1,0),(+1,0),(0,-1),(0,+1),(-1,-1),(-1,+1)]
NEIGHBORS_ODD  = [(-1,0),(+1,0),(0,-1),(0,+1),(+1,-1),(+1,+1)]

def open_sides(y, x, tiles, Y, X):
    offsets = NEIGHBORS_EVEN if y%2==0 else NEIGHBORS_ODD
    return sum(1 for dy,dx in offsets
               if 0<=y+dy<Y and 0<=x+dx<X
               and tiles[y+dy][x+dx].get('TileType',1)!=1)

def color_changes(stacks):
    return sum(1 for i in range(1,len(stacks)) if stacks[i]!=stacks[i-1])

def analyze_level(data):
    Y,X = data['YCells'], data['XCells']
    tiles = data['Tiles']
    groups = {t:[] for t in TILETYPE.values()}
    for y in range(Y):
        for x in range(X):
            t  = tiles[y][x]
            tt = t.get('TileType',0)
            groups[TILETYPE[tt]].append((y,x,t))
    def ss(lst): return sum(open_sides(y,x,tiles,Y,X) for y,x,_ in lst)
    sa = groups['Stack']+groups['StackLock']
    return {
        'XCells':X,'YCells':Y,
        'H1_1':X*Y, 'H1_2':ss(groups['Normal']),
        'H1_3':len(groups['Normal']), 'H1_4':ss(sa),
        'H1_5':len(sa),
        'H1_6':sum(len(t.get('Stacks',[])) for _,_,t in sa),
        'H1_7':sum(color_changes(t.get('Stacks',[])) for _,_,t in sa),
        'H1_8':ss(groups['Lock']), 'H1_9':len(groups['Lock']),
        'H1_10':ss(groups['StackLock']), 'H1_11':len(groups['StackLock']),
        'H1_12':(sum(t.get('Level',0) for _,_,t in groups['Lock']+groups['Plank'])+
                 sum(t.get('UnlockLevel',0) for _,_,t in groups['StackLock']+groups['Ice'])),
        'H1_13':ss(groups['Ads']), 'H1_14':min(len(groups['Ads']),3),
        'H1_15':ss(groups['Plank']+groups['Ice']+groups['Grass']+groups['CameraPicture']),
        'tile_counts':{k:len(v) for k,v in groups.items()},
    }

# ── 난이도 공식
def baseline_curve(n_arr):
    return 70 - 52*np.exp(-n_arr/90)

LOCAL_VAR_SEED = [
    3,-2,5,-3,4,-1,6,-4,3,-2,5,-1,4,-3,2,-5,6,-2,3,-4,
    5,-1,4,-3,2,-6,5,-2,4,-1,3,-5,6,-3,2,-4,5,-1,4,-2,
    3,-6,5,-3,4,-1,2,-5,6,-4,3,-2,5,-1,4,-3,6,-2,3,-5,
    4,-1,5,-3,2,-6,4,-2,5,-1,3,-4,6,-2,4,-5,3,-1,5,-3,
    4,-2,6,-1,3,-5,4,-2,5,-3,4,-1,6,-4,2,-5,3,-2,5,-1,
]
def target_curve(n_arr):
    base = baseline_curve(n_arr)
    lv = np.array([LOCAL_VAR_SEED[(int(n)-1)%100] for n in n_arr], dtype=float)
    return np.clip(base + 3.71 + lv, 0, 100)

# ── GitHub 커밋
def github_commit(token, repo, path, content_bytes, message):
    url = f"https://api.github.com/repos/{repo}/contents/{path}"
    headers = {"Authorization": f"token {token}", "Accept": "application/vnd.github.v3+json"}
    r = requests.get(url, headers=headers)
    sha = r.json().get("sha") if r.status_code==200 else None
    body = {"message": message, "content": base64.b64encode(content_bytes).decode()}
    if sha: body["sha"] = sha
    resp = requests.put(url, headers=headers, json=body)
    return resp.status_code in (200, 201), resp.json()

# ══════════════════════════════════════════════════════
# 캐시 로더
# ══════════════════════════════════════════════════════
@st.cache_data
def load_intg_local():
    return pd.read_csv(INTG_CSV) if INTG_CSV.exists() else pd.DataFrame()

@st.cache_data
def load_tbl_local():
    """GitHub 저장소의 tblStage_500.xlsx에서 N_ 레벨 행만 로드."""
    if not TBLSTAGE.exists(): return pd.DataFrame()
    df = pd.read_excel(TBLSTAGE, sheet_name='Stage', header=0)
    return df[df['LevelName'].str.startswith('N ', na=False)].reset_index(drop=True)

@st.cache_data
def load_level_local(lv):
    """로컬 data/levels/에서 N_{lv:03d}.json 파일을 로드."""
    p = LEVELS_DIR/f"N_{lv:03d}.json"
    return json.load(open(p)) if p.exists() else None

# ══════════════════════════════════════════════════════
# 사이드바
# ══════════════════════════════════════════════════════
LANG = st.session_state.get("lang","한국어")
IS_EN = (LANG == "English")

with st.sidebar:
    st.markdown("""
<div style="text-align:center;padding:8px 0 4px 0;">
  <span style="font-size:1.6em;font-weight:800;color:#6B3A2A;">🧩 Puzzle Creator</span><br>
  <span style="font-size:0.75em;color:#7A5C45;">My Little Bookstore</span>
</div>
""", unsafe_allow_html=True)
    st.markdown("---")

    # ── 페이지 선택
    st.markdown(f'<div class="sidebar-label">{TR("page")}</div>', unsafe_allow_html=True)
    if IS_EN:
        page_options = ["🏠 Home","📖 1. Manual","📊 2. Difficulty Analysis",
                        "🗺️ 3. Board Viewer","🎲 4. JSON Generator",
                        "🔧 5. Settings","🗄️ 6. Archive",
                        "🧩 7. Special Puzzle"]
    else:
        page_options = ["🏠 홈","📖 1. 매뉴얼","📊 2. 난이도 분석",
                        "🗺️ 3. 판 모양 뷰어","🎲 4. JSON 생성기",
                        "🔧 5. 설정","🗄️ 6. 아카이브",
                        "🧩 7. 묘수풀이 생성기"]
    # 언어 바뀌어도 같은 탭 유지
    PAGE_MAP_KR = ["🏠 홈","📖 1. 매뉴얼","📊 2. 난이도 분석",
                   "🗺️ 3. 판 모양 뷰어","🎲 4. JSON 생성기","🔧 5. 설정","🗄️ 6. 아카이브",
                   "🧩 7. 묘수풀이 생성기"]
    PAGE_MAP_EN = ["🏠 Home","📖 1. Manual","📊 2. Difficulty Analysis",
                   "🗺️ 3. Board Viewer","🎲 4. JSON Generator","🔧 5. Settings","🗄️ 6. Archive",
                   "🧩 7. Special Puzzle"]
    cur_page_key = st.session_state.get("cur_page_idx", 0)
    page_idx = st.radio("", range(len(page_options)),
                        format_func=lambda i: page_options[i],
                        index=cur_page_key,
                        label_visibility="collapsed")
    st.session_state["cur_page_idx"] = page_idx
    page_kr = PAGE_MAP_KR[page_idx]
    page = page_kr  # 내부 조건문은 한국어 기준 유지

    st.markdown("---")

    # ── 파일 업로드
    st.markdown(f'<div class="sidebar-label">{TR("file_upload")}</div>', unsafe_allow_html=True)

    up_intg = st.file_uploader("integrated_difficulty.csv", type=["csv"], key="up_intg")
    if up_intg:
        st.session_state.intg_df = pd.read_csv(up_intg)
        st.markdown('<span class="file-badge">✓ 통합 난이도</span>', unsafe_allow_html=True)
    elif st.session_state.intg_df is None:
        local = load_intg_local()
        if not local.empty:
            st.session_state.intg_df = local
            st.markdown('<span class="file-badge">✓ 통합 난이도 (로컬)</span>', unsafe_allow_html=True)
        else:
            st.markdown('<span class="file-badge-warn">⚠ 통합 난이도 없음</span>', unsafe_allow_html=True)

    up_tbl = st.file_uploader("tblStage_500.xlsx", type=["xlsx"], key="up_tbl")
    if up_tbl:
        df_raw = pd.read_excel(up_tbl, sheet_name='Stage', header=0)
        st.session_state.tbl_df = df_raw[df_raw['LevelName'].str.startswith('N ', na=False)].reset_index(drop=True)
        st.markdown('<span class="file-badge">✓ tblStage</span>', unsafe_allow_html=True)
    elif st.session_state.tbl_df is None:
        local = load_tbl_local()
        if not local.empty:
            st.session_state.tbl_df = local
            st.markdown('<span class="file-badge">✓ tblStage (로컬)</span>', unsafe_allow_html=True)
        else:
            st.markdown('<span class="file-badge-warn">⚠ tblStage 없음</span>', unsafe_allow_html=True)

    up_market = st.file_uploader(TR("market_upload"), type=["csv"], key="up_market")
    if up_market:
        raw = pd.read_csv(up_market, header=0, skiprows=[1,2])
        raw.columns = raw.columns.str.strip()
        st.session_state.market_df = raw
        st.session_state.market_source = "업로드"
        st.markdown('<span class="file-badge">✓ 시장 데이터 (업로드)</span>', unsafe_allow_html=True)
    elif st.session_state.market_df is not None:
        src = st.session_state.get("market_source", "로컬")
        st.markdown(f'<span class="file-badge">✓ 시장 데이터 ({src})</span>', unsafe_allow_html=True)
    elif MARKET_CSV.exists():
        try:
            raw = pd.read_csv(MARKET_CSV, header=0, skiprows=[1, 2])
            raw.columns = raw.columns.str.strip()
            st.session_state.market_df = raw
            st.session_state.market_source = "기본 내장"
            st.markdown('<span class="file-badge">✓ Market Data (Built-in)" if IS_EN else "시장 데이터 (기본 내장)</span>', unsafe_allow_html=True)
        except Exception:
            st.markdown('<span class="file-badge-warn">⚠ 시장 데이터 로드 실패</span>', unsafe_allow_html=True)
    else:
        st.markdown('<span class="file-badge-warn">⚠ 시장 데이터 없음</span>', unsafe_allow_html=True)

    st.markdown("---")

    # ── 다운로드
    st.markdown(f'<div class="sidebar-label">{TR("download")}</div>', unsafe_allow_html=True)
    intg = st.session_state.intg_df
    if intg is not None:
        st.download_button(TR("csv_intg"), df_to_csv_bytes(intg),
                           "integrated_difficulty.csv", "text/csv", use_container_width=True)
        st.download_button(TR("json_intg"), df_to_json_bytes(intg),
                           "integrated_difficulty.json", "application/json", use_container_width=True)
    else:
        st.caption(TR("upload_hint"))

    st.markdown("---")

    # ── GitHub 토큰
    with st.expander(TR("github_settings")):
        token_input = st.text_input("Personal Access Token", type="password",
                                    value=st.session_state.github_token)
        if token_input:
            st.session_state.github_token = token_input
        st.caption(f"Repo: {GITHUB_REPO}")



    st.markdown("---")
    # ── 언어 (후순위)
    st.markdown('<div class="sidebar-label">🌐 Language</div>', unsafe_allow_html=True)
    lang_sel = st.radio(
        "🌐 Language",
        ["한국어", "English"],
        index=0 if st.session_state.get("lang","한국어")=="한국어" else 1,
        horizontal=True,
        key="lang_radio"
    )
    if lang_sel != st.session_state.get("lang","한국어"):
        st.session_state["lang"] = lang_sel
        st.rerun()
    # 사이드바 내에서도 IS_EN 갱신
    LANG = st.session_state.get("lang","한국어")
    IS_EN = (LANG == "English")

# ══════════════════════════════════════════════════════
# 탭 0 — 홈
# ══════════════════════════════════════════════════════
if page == "🏠 홈":
    GITHUB_RAW_BASE = f"https://raw.githubusercontent.com/{GITHUB_REPO}/main"

    # ── 페이지 애니메이션 wrapper
    st.markdown('<div class="page-anim">', unsafe_allow_html=True)

    # ── 이미지 base64 로더
    def img_to_b64(filename):
        """assets/images/ 경로의 이미지를 base64 문자열로 변환."""
        p = BASE / "assets" / "images" / filename
        if p.exists():
            import base64 as _b64
            return "data:image/png;base64," + _b64.b64encode(p.read_bytes()).decode()
        return f"{GITHUB_RAW_BASE}/assets/images/{quote(filename)}"

    bg_src    = img_to_b64("BG.png")
    icon_src  = img_to_b64("Icon.png")
    logo_src  = img_to_b64("Logo.png")

    # ── 히어로 섹션
    _hero_sub = TR("hero_sub")
    st.markdown(f"""
<div style="position:relative;border-radius:20px;overflow:hidden;margin-bottom:24px;
            box-shadow:0 8px 32px rgba(107,58,42,0.20);">
  <img src="{bg_src}" style="width:100%;height:260px;object-fit:cover;display:block;">
  <div style="position:absolute;top:0;left:0;right:0;bottom:0;
              background:linear-gradient(120deg,rgba(44,24,16,0.55) 0%,rgba(107,58,42,0.25) 100%);
              display:flex;align-items:center;padding:28px 36px;gap:24px;">
    <img src="{icon_src}" style="height:88px;border-radius:18px;
         box-shadow:0 4px 16px rgba(0,0,0,0.35);flex-shrink:0;">
    <div style="flex:1;">
      <img src="{logo_src}" style="height:80px;display:block;margin-bottom:10px;
           filter:drop-shadow(0 2px 6px rgba(0,0,0,0.4));">
      <p style="color:#FBF5EE;font-size:13px;margin:10px 0 0 0;
                text-shadow:0 1px 3px rgba(0,0,0,0.5);">{_hero_sub}</p>
    </div>
  </div>
</div>
""", unsafe_allow_html=True)

    # 핵심 지표
    m1, m2, m3, m4 = st.columns(4)
    m1.metric(TR("metric_levels"), "500")
    m2.metric(TR("metric_market"), "Lv 1~100")
    m3.metric(TR("metric_metrics"), "15 (H1)")
    m4.metric(TR("metric_launch"), "2026. 10")
    st.markdown('<hr style="border-color:#E8D5C0;margin:24px 0;">', unsafe_allow_html=True)


    # ── 게임 플레이 영상
    st.markdown(f'<div class="section-header">{TR("sec_video")}</div>', unsafe_allow_html=True)
    video_path = BASE / "assets" / "videos" / "Hexasort Puzzle.mp4"
    video_url  = f"{GITHUB_RAW_BASE}/assets/videos/{quote('Hexasort Puzzle.mp4')}"
    col_vid, col_pad = st.columns([1, 2])
    with col_vid:
        if video_path.exists():
            with open(video_path, "rb") as vf:
                st.video(vf.read())
        else:
            st.video(video_url)

    st.markdown('<hr style="border-color:#E8D5C0;margin:24px 0;">', unsafe_allow_html=True)

    # ── 게임 소개 슬라이드 (18장 애니메이션)
    st.markdown(f'<div class="section-header">{TR("sec_intro")}</div>', unsafe_allow_html=True)
    st.markdown("""
<style>
@keyframes cardIn {
    from { opacity:0; transform:translateY(32px) scale(0.97); }
    to   { opacity:1; transform:translateY(0) scale(1); }
}
.intro-card {
    background:#FFFFFF; border-radius:14px;
    box-shadow:0 4px 20px rgba(107,58,42,0.12);
    overflow:hidden; margin-bottom:20px;
    opacity:0;
    animation:cardIn 0.5s ease forwards;
}
.intro-card img { width:100%; display:block; }
.intro-num {
    font-size:12px; color:#7A5C45; font-weight:600;
    padding:6px 12px; background:#F0E6D8; text-align:right;
}
</style>
""", unsafe_allow_html=True)

    # 2열로 18장 표시
    imgs = [f"{str(i).zfill(2)}.png" for i in range(1, 19)]
    # 존재하는 파일만
    img_paths = [(BASE / "assets" / "images" / img) for img in imgs]

    for row_i in range(0, 18, 2):
        cols = st.columns(2)
        for col_j, c in enumerate(cols):
            idx = row_i + col_j
            if idx >= 18: break
            img_name = imgs[idx]
            img_path = img_paths[idx]
            delay = f"{0.1 + idx*0.07:.2f}s"
            img_src = img_to_b64(img_name)
            with c:
                st.markdown(
                    f'''<div class="intro-card" style="animation-delay:{delay};">
                    <div class="intro-num">{idx+1} / 18</div>
                    <img src="{img_src}" alt="게임 소개 {idx+1}"
                         style="width:100%;display:block;border-radius:8px;">
                    </div>''',
                    unsafe_allow_html=True
                )

    st.markdown('</div>', unsafe_allow_html=True)  # page-anim 닫기

# ══════════════════════════════════════════════════════
# 탭 1 — 매뉴얼
# ══════════════════════════════════════════════════════
if page == "📖 1. 매뉴얼":
    st.title("📖 Puzzle Creator 사용 매뉴얼" if not IS_EN else "📖 Puzzle Creator User Manual")
    st.caption("My Little Bookstore — 헥사소트 퍼즐 레벨 난이도 설계 & 분석 도구" if not IS_EN else "My Little Bookstore — Hexasort Puzzle Level Difficulty Design & Analysis Tool")

    st.markdown("---")

    # ── 앱 목표 섹션 (한/영)
    if not IS_EN:
        st.markdown("## 🎯 이 앱의 목표")
        st.markdown("""
**Puzzle Creator**는 헥사소트 퍼즐 게임 *나의 작은 서점 (My Little Bookstore)*의
레벨 난이도를 **데이터 기반으로 설계하고 자동 생성**하는 통합 도구입니다.

### 배경 — 헥사소트 퍼즐이란?
헥사소트(Hexasort)는 육각형 타일에 쌓인 색상 칩을 같은 색끼리 정렬하는 모바일 퍼즐 게임입니다.
난이도는 두 가지 요소로 결정됩니다:
- **초기 판 모양**: 타일 수, 스택 위치, 잠금 타일, 기믹 배치 등
- **게임 진행 파라미터**: 할당량, 색상 풀, 중복률, 색상 추가 시점 등

### 왜 만들었나요?
| 문제 | 기존 방식 | Puzzle Creator |
|---|---|---|
| 500개 레벨 생성 | PPT 수작업 (작업자 직접) | 공식 기반 자동 생성 |
| 난이도 기준 없음 | 주관적 판단 | 시장 데이터 기반 객관화 |
| 판 확인 불가 | JSON 파일 직접 열어봄 | 헥사 그리드 시각화 |
| 가중치 조정 어려움 | 코드 수정 후 재실행 | 슬라이더 실시간 조정 |

### 어떻게 객관화했나요?
SKKU 게임센터 랩에서 시장 1위 헥사소트 게임의 **Lv 1~100을 직접 플레이하며 판 구성 데이터를 수집**했습니다.
이 실측 데이터에서 **H1-1~H1-15 총 15개 지표**를 정의하고, 이를 정규화·가중치 합산하여
레벨별 난이도 점수(0~100)를 계산한 뒤 **지수수렴 난이도 곡선 공식**을 도출했습니다.
        """)
    else:
        st.markdown("## 🎯 What This App Does")
        st.markdown("""
**Puzzle Creator** is an integrated tool for **data-driven level difficulty design and automated generation**
for the hexasort puzzle game *My Little Bookstore*.

### Background — What is Hexasort?
Hexasort is a mobile puzzle game where players sort colored chips stacked on hexagonal tiles by matching colors.
Difficulty is determined by two components:
- **Initial board layout**: tile count, stack positions, locked tiles, gimmick placement, etc.
- **Gameplay parameters**: allocation target, color pool, duplication rate, color unlock timing, etc.

### Why was this built?
| Problem | Previous Approach | Puzzle Creator |
|---|---|---|
| Generate 500 levels | Manual PowerPoint work | Auto-generation via formula |
| No difficulty standard | Subjective judgment | Objectified via market data |
| Can't preview boards | Open JSON files manually | Hex grid visualization |
| Hard to tune weights | Edit code & re-run | Real-time slider adjustment |

### How was it objectified?
The SKKU Game Center Lab **directly played and collected board data from Levels 1–100** of the
top-ranked hexasort game on the market. From this real-world dataset, **15 indicators (H1-1~H1-15)**
were defined, normalized, and weighted to compute a difficulty score (0–100) per level,
from which an **exponential convergence difficulty curve formula** was derived.
        """)

    st.markdown("---")
    if not IS_EN:
        st.markdown("## 전체 데이터 흐름")
    else:
        st.markdown("## Data Pipeline")

    if not IS_EN:
        st.code("""
① 시장 데이터 분석
   market_lv1_100.csv (SKKU 게임센터 랩 실측, Lv 1~100)
        | H1-1~H1-15 지표 추출 + 가중치 합산
   board_score 계산 -> 난이도 곡선 패턴 도출
        |
   target(N) = (70 - 52 x e^(-N/90)) + 3.71 + local_var[(N-1) mod 100]

② JSON 판 모양 생성 (generate_levels.py)
   target(N) 목표 설정
        | 판 반복 생성 (최대 10회)
   board_score ≈ target(N) 되는 판 선택
        |
   stack_score_target = target(N) x 2 - board_score
        | 역산
   tblStage 파라미터 자동 계산
        |
   통합 난이도 = (board_score + stack_score) / 2 ≈ target(N)

③ 시각화 & 검증 (이 앱)
   JSON + tblStage -> 난이도 곡선 비교 / 판 모양 확인 / 재생성
        """, language="")
    else:
        st.code("""
① Market Data Analysis
   market_lv1_100.csv (SKKU Game Center Lab, real-world Lv 1~100)
        | Extract H1-1~H1-15 indicators + weighted sum
   Compute board_score -> derive difficulty curve pattern
        |
   target(N) = (70 - 52 x e^(-N/90)) + 3.71 + local_var[(N-1) mod 100]

② JSON Board Generation (generate_levels.py)
   Set target(N) as difficulty goal
        | Retry board generation (max 10 times)
   Select board where board_score ≈ target(N)
        |
   stack_score_target = target(N) x 2 - board_score
        | Inverse calculation
   Auto-compute tblStage parameters
        |
   Integrated difficulty = (board_score + stack_score) / 2 ≈ target(N)

③ Visualization & Validation (this app)
   JSON + tblStage -> Compare curves / Preview boards / Regenerate
        """, language="")

    st.markdown("---")
    st.markdown(TR("tab_guide"))
    cols = st.columns(2)
    tabs_info = [
        ("📖 1. Manual" if IS_EN else "📖 1. 매뉴얼", "Overview, glossary, and file structure." if IS_EN else "지금 보고 계신 페이지입니다. 전체 흐름, 용어, 파일 구조를 안내합니다."),
        ("📊 2. Difficulty Analysis" if IS_EN else "📊 2. 난이도 분석", "Market data + integrated difficulty curve. Adjust board/gameplay weights." if IS_EN else "시장 데이터 원본 테이블 + 우리 게임 통합 난이도 곡선 비교. 판모양/게임진행 가중치 조정 가능."),
        ("🗺️ 3. Board Viewer" if IS_EN else "🗺️ 3. 판 모양 뷰어", "Visualize JSON as hex grid. Edit tiles, save JSON, live difficulty score." if IS_EN else "레벨 JSON을 헥사 그리드로 시각화. 편집 모드에서 타일 수정 후 JSON 저장. 실시간 난이도 점수 표시."),
        ("🎲 4. JSON Generator" if IS_EN else "🎲 4. JSON 생성기", "Select range → preview curve → generate → download ZIP." if IS_EN else "레벨 범위 입력 -> 난이도 곡선 미리보기 -> 생성 -> 실시간 진행 표시 -> zip 다운로드."),
        ("🔧 5. Settings" if IS_EN else "🔧 5. 설정", "Adjust H1 and tblStage weights. Edit stack parameters." if IS_EN else "H1 지표별 가중치(%) 조정 및 tblStage 스택 파라미터 직접 수정."),
        ("🗄️ 6. Archive" if IS_EN else "🗄️ 6. 아카이브", "Save settings as versions, auto-commit to GitHub." if IS_EN else "설정값을 버전으로 저장하고 GitHub에 자동 커밋. 버전 간 비교 가능."),
    ]
    for i, (name, desc) in enumerate(tabs_info):
        with cols[i%2]:
            st.markdown(f"""
<div class="metric-card" style="text-align:left; margin-bottom:12px;">
<strong>{name}</strong><br>
<span style="font-size:13px;color:#8b949e;">{desc}</span>
</div>""", unsafe_allow_html=True)

    st.markdown("---")
    st.markdown(TR("file_structure"))
    if IS_EN:
        st.markdown("""
| File | Location | Description |
|---|---|---|
| `market_lv1_100.csv` | `data/market/` | Market game Lv 1~100 H1 data (built-in) |
| `tblStage_500.xlsx` | `data/` | 500-stage gameplay parameters |
| `N_001.json` ~ `N_500.json` | `data/levels/` | Board layout JSON per level |
| `integrated_difficulty.csv` | `data/` | Integrated difficulty scores per level |
| `generate_levels.py` | root | Auto-generation script for JSON + tblStage |
| `level_analyzer_v2.py` | root | JSON → H1 indicator extractor |

> Market data CSV is built-in — no upload needed.
        """)
    else:
        st.markdown("""
| 파일 | 위치 | 설명 |
|---|---|---|
| `market_lv1_100.csv` | `data/market/` | 시장 게임 Lv 1~100 H1 실측값 (기본 내장) |
| `tblStage_500.xlsx` | `data/` | 500개 스테이지 게임 파라미터 |
| `N_001.json` ~ `N_500.json` | `data/levels/` | 레벨별 판 모양 JSON |
| `integrated_difficulty.csv` | `data/` | 레벨별 통합 난이도 |
| `generate_levels.py` | 루트 | JSON + tblStage 자동 생성 스크립트 |
| `level_analyzer_v2.py` | 루트 | JSON → H1 지표 추출 파서 |

> 시장 데이터 CSV는 앱에 기본 내장되어 있어 별도 업로드 없이 사용 가능합니다.
        """)

    st.markdown("---")
    st.markdown(TR("update_json"))
    if IS_EN:
        st.markdown("""
**Method A — Generate in App (Recommended)**
1. Go to `🎲 4. JSON Generator` tab
2. Select level range with slider
3. Preview difficulty curve
4. Click **Generate** → check real-time difficulty per level
5. **Download ZIP**
6. Upload to GitHub `data/levels/` (overwrite)

**Method B — Edit Individual Levels**
1. `🗺️ 3. Board Viewer` → select level
2. Turn on Edit Mode → modify tiles
3. Save JSON → replace only that file on GitHub
        """)
    else:
        st.markdown("""
**방법 A — 앱에서 직접 생성 (추천)**
1. `🎲 4. JSON 생성기` 탭으로 이동
2. 레벨 범위 슬라이더로 생성할 구간 선택
3. 난이도 곡선 미리보기 확인
4. **생성** 버튼 클릭 → 실시간으로 레벨별 난이도 점수 확인
5. 완료 후 **zip 다운로드**
6. GitHub `data/levels/` 폴더에 덮어쓰기 업로드

**방법 B — 개별 레벨 수정**
1. `🗺️ 3. 판 모양 뷰어` → 레벨 선택
2. 편집 모드 ON → 타일 수정
3. JSON 저장 → GitHub에 해당 파일만 교체
        """)

    st.markdown("---")
    st.markdown(TR("glossary"))

    with st.expander(TR("tiletype_exp")):
        if IS_EN:
            st.markdown("""
| Code | Name | Description | Unlock Level |
|---|---|---|---|
| 0 | Normal | Empty cell where players place stacks | Lv 1~ |
| 1 | Blank | Inactive cell (outside grid) | — |
| 2 | Stack | Cell with pre-stacked chips | Lv 1~ |
| 3 | Lock | Locked cell — unlocked by Level score | Lv 9~ |
| 4 | Plank | Wood plank — destroyed by Level score | Lv 59~ |
| 5 | Ice | Ice — unlocked by UnlockLevel, includes stack | Lv 179~ |
| 6 | StackLock | Locked stack — unlocked by UnlockLevel | Lv 29~ |
| 7 | Grass | Grass — removed when sorted from above | Lv 299~ |
| 8 | Ads | Ad tile — watch ad to gain empty cell | Lv 49~ |
            """)
        else:
            st.markdown("""
| 코드 | 이름 | 설명 | 등장 레벨 |
|---|---|---|---|
| 0 | Normal | 플레이어가 스택을 놓는 일반 빈 칸 | Lv 1~ |
| 1 | Blank | 비활성 셀 (그리드 밖) | — |
| 2 | Stack | 초기부터 칩이 쌓인 셀 | Lv 1~ |
| 3 | Lock | 잠긴 셀 — Level 달성 시 해제 | Lv 9~ |
| 4 | Plank | 나무판 — Level 달성 시 파괴 | Lv 59~ |
| 5 | Ice | 얼음 — UnlockLevel 달성 시 해제, 스택 포함 | Lv 179~ |
| 6 | StackLock | 잠긴 스택 — UnlockLevel 달성 시 해제 | Lv 29~ |
| 7 | Grass | 잔디 — 위에서 소팅 시 제거 | Lv 299~ |
| 8 | Ads | 광고 — 시청 시 빈 칸 획득 | Lv 49~ |
            """)

    with st.expander(TR("h1_exp")):
        if IS_EN:
            st.markdown("""
Normalized by market data min/max, then weighted sum → board_score (0~100)

| Indicator | Description | Difficulty Direction |
|---|---|---|
| H1-1 | Total grid count | More = Easier |
| H1-2 | Normal cell open-side sum | Lower = Harder |
| H1-3 | Normal cell count | Fewer = Harder |
| H1-4 | Stack+StackLock open-side sum | Lower = Harder |
| H1-5 | Stack+StackLock count | More = Harder |
| H1-6 | Total chip colors | More = Harder |
| H1-7 | Color change complexity | More = Harder |
| H1-8 | Lock open-side sum | More = Harder |
| H1-9 | Lock count | More = Harder |
| H1-10 | StackLock open-side sum | More = Harder |
| H1-11 | StackLock count | More = Harder |
| H1-12 | Unlock threshold sum (Level + UnlockLevel) | Higher = Harder |
| H1-13 | Ads open-side sum | Lower = Harder |
| H1-14 | Ads count (max 3) | More = Easier |
| H1-15 | Gimmick open-side sum (Plank/Ice/Grass/Camera) | Lower = Harder |
            """)
        else:
            st.markdown("""
시장 데이터 min/max로 정규화 후 가중치 합산 → board_score (0~100점)

| 지표 | 설명 | 난이도 방향 |
|---|---|---|
| H1-1 | 전체 그리드 수 | 많을수록 쉬움 |
| H1-2 | Normal 셀 열린 변 합 | 낮을수록 어려움 |
| H1-3 | Normal 셀 개수 | 적을수록 어려움 |
| H1-4 | Stack+StackLock 열린 변 합 | 낮을수록 어려움 |
| H1-5 | Stack+StackLock 개수 | 많을수록 어려움 |
| H1-6 | 타일 색 총합 | 많을수록 어려움 |
| H1-7 | 색 변화 복잡도 | 많을수록 어려움 |
| H1-8 | Lock 열린 변 합 | 많을수록 어려움 |
| H1-9 | Lock 개수 | 많을수록 어려움 |
| H1-10 | StackLock 열린 변 합 | 많을수록 어려움 |
| H1-11 | StackLock 개수 | 많을수록 어려움 |
| H1-12 | 잠금 해제 기준 합 (Level + UnlockLevel) | 클수록 어려움 |
| H1-13 | Ads 열린 변 합 | 낮을수록 어려움 |
| H1-14 | Ads 수 (최대 3 캡) | 많을수록 쉬움 |
| H1-15 | 기믹 열린 변 합 (Plank/Ice/Grass/Camera) | 낮을수록 어려움 |
            """)

    with st.expander(TR("tbl_exp")):
        if IS_EN:
            st.markdown("""
| Parameter | Description |
|---|---|
| TotalAllocation | Target score (allocation). Reaching it triggers next color unlock |
| InitialAvailableColors | Color pool at game start. **Top chip on initial board uses only these colors** |
| DistinctColorCount | Number of colors per stack. Updated at each threshold |
| ColorDuplicationRate | Probability of generating identical color stacks. Lower = Harder |
| ProgressAddNewColor | Color unlock thresholds. Evenly distributed within TotalAllocation. 1:1 with NewColorsMilestones |
| NewColorsMilestones | Colors added at each threshold. Cannot overlap with InitialAvailableColors |
            """)
        else:
            st.markdown("""
| 파라미터 | 설명 |
|---|---|
| TotalAllocation | 목표 점수(할당량). 채우면 다음 색상 추가 임계값 도달 |
| InitialAvailableColors | 게임 시작 시 사용 색상 풀. **초기 판의 맨 위 칩은 이 색상만 사용** |
| DistinctColorCount | 스택 하나에서 사용하는 색상 수. 임계값마다 업데이트 |
| ColorDuplicationRate | 완전히 같은 색상 스택 생성 확률. 낮을수록 어려움 |
| ProgressAddNewColor | 색상 추가 임계값. TotalAllocation 내 균등 분배. NewColorsMilestones와 1:1 |
| NewColorsMilestones | 임계값 도달 시 추가되는 색상. InitialAvailableColors와 중복 불가 |
            """)

    with st.expander(TR("chip_exp")):
        cols2 = st.columns(4)
        chip_info = [(0,'🔵 Blue','#1890FF'),(1,'🟡 Yellow','#FADB14'),
                     (2,'🔴 Red','#F5222D'),(3,'🟢 Green','#52C41A'),
                     (4,'🟠 Orange','#FA8C16'),(5,'🟣 Purple','#722ED1'),
                     (6,'⬜ White','#AAAAAA'),(7,'⬛ Black','#333333')]
        for i,(code,name,_) in enumerate(chip_info):
            cols2[i%4].markdown(f"**{code}** — {name}")

    with st.expander(TR("grade_exp")):
        if IS_EN:
            st.markdown("""
| Score | Grade | Color |
|---|---|---|
| 0~25 | Very Easy | 🔵 Blue |
| 25~45 | Easy | 🟢 Green |
| 45~60 | Normal | 🟡 Yellow |
| 60~75 | Hard | 🟠 Orange |
| 75+ | Very Hard | 🔴 Red |
            """)
        else:
            st.markdown("""
| 점수 | 등급 | 색상 |
|---|---|---|
| 0~25 | 매우쉬움 | 🔵 파랑 |
| 25~45 | 쉬움 | 🟢 초록 |
| 45~60 | 보통 | 🟡 노랑 |
| 60~75 | 어려움 | 🟠 주황 |
| 75+ | 매우어려움 | 🔴 빨강 |
            """)

    with st.expander(TR("formula_exp")):
        if IS_EN:
            st.markdown("""
**Derived from real market game Lv 1~100 data (SKKU Game Center Lab)**

```
target(N) = (70 - 52 x e^(-N/90)) + 3.71 + local_var[(N-1) mod 100]
```

- **Base curve**: Lv1 ≈ 18pt → Lv100 ≈ 48pt → converges to ~74pt
- **local_var**: 100-value oscillation pattern from market data (repeats every 100 levels)
- **Integrated difficulty**: board_score x 50% + stack_score x 50% ≈ target(N)
            """)
        else:
            st.markdown("""
**시장 게임 Lv 1~100 데이터 분석으로 도출 (SKKU 게임센터 랩)**

```
target(N) = (70 - 52 x e^(-N/90)) + 3.71 + local_var[(N-1) mod 100]
```

- **기준선**: Lv1≈18점 → Lv100≈48점 → 최대 ~74점 수렴
- **local_var**: 시장 데이터에서 추출한 100개 오르내림 패턴 (100레벨 주기 반복)
- **통합 난이도**: board_score x 50% + stack_score x 50% ≈ target(N)
            """)


# ══════════════════════════════════════════════════════
# 탭 2 — 난이도 분석
# ══════════════════════════════════════════════════════
elif page == "📊 2. 난이도 분석":
    st.markdown('<div class="page-anim">', unsafe_allow_html=True)
    st.title(TR("analysis_title"))
    st.caption(TR("analysis_cap"))

    intg   = st.session_state.intg_df
    market = st.session_state.market_df

    # ── 시장 데이터 원본 테이블 (곡선 위쪽)
    with st.expander(TR("market_exp"), expanded=False):
        if market is not None:
            src = st.session_state.get("market_source", "로컬")
            st.caption(f"SKKU Game Center Lab real data · {len(market)} levels · Source: {src}" if IS_EN else f"SKKU 게임센터 랩 실측 데이터 · {len(market)}개 레벨 · 출처: {src}")
            # 원본 테이블
            mk_show = market.copy()
            if "Stage" in mk_show.columns:
                mk_show = mk_show[mk_show["Stage"].apply(
                    lambda x: str(x).strip().lstrip("-").isdigit() if pd.notna(x) else False)]
            st.dataframe(mk_show.reset_index(drop=True), use_container_width=True, height=300)
            # H1 지표 추이 차트
            st.markdown(TR("h1_trend"))
            mk_cols_norm = {c.strip(): c for c in mk_show.columns}
            h1_avail = [mk_cols_norm[c] for c in
                        ["H1-1","H1-2","H1-3","H1-4","H1-5",
                         "H1-6","H1-7","H1-8","H1-9","H1-12","H1-13","H1-14"]
                        if c in mk_cols_norm]
            sel_h1 = st.multiselect(TR("show_indicators"), h1_avail,
                                     default=h1_avail[:4], key="market_h1_sel")
            if sel_h1 and "Stage" in mk_show.columns:
                mk_valid = mk_show.copy()
                mk_valid["Stage"] = mk_valid["Stage"].astype(int)
                fig_mk = go.Figure()
                for col in sel_h1:
                    vals = pd.to_numeric(mk_valid[col], errors="coerce")
                    fig_mk.add_trace(go.Scatter(
                        x=mk_valid["Stage"], y=vals,
                        name=col.strip(), mode="lines+markers", marker=dict(size=4)
                    ))
                fig_mk.update_layout(
                    height=300,
                    plot_bgcolor=T["plot_bg"], paper_bgcolor=T["plot_bg"],
                    font_color=T["text"], xaxis_title="Level" if IS_EN else "Level" if IS_EN else "레벨", yaxis_title="값",
                    xaxis=dict(gridcolor=T["grid_line"]),
                    yaxis=dict(gridcolor=T["grid_line"]),
                    legend=dict(orientation="h", y=1.12, bgcolor="rgba(0,0,0,0)"),
                    margin=dict(l=10,r=10,t=30,b=10)
                )
                st.plotly_chart(fig_mk, use_container_width=True)
        else:
            st.info("사이드바에서 시장 데이터 CSV를 업로드하거나, data/market/ 폴더에 market_lv1_100.csv를 추가하세요.")

    st.markdown("---")

    # ── 가중치 슬라이더 (차트 위)
    st.subheader(TR("weight_adj"))
    if intg is not None:
        w_b = st.slider(TR("board_weight"), 0, 100, st.session_state.w_board, key="sl_wboard")
        w_g = 100 - w_b
        st.session_state.w_board    = w_b
        st.session_state.w_gameplay = w_g
        st.caption(f"Board **{w_b}%** : Gameplay **{w_g}%**" if IS_EN else f"판 모양 **{w_b}%** : 게임 진행 **{w_g}%**")
    else:
        w_b = st.session_state.w_board
        w_g = st.session_state.w_gameplay
        st.info("integrated_difficulty.csv를 사이드바에서 업로드하면 가중치를 조정할 수 있습니다.")

    # ── 차트 표시 옵션
    st.markdown(TR("select_curves"))
    oc1, oc2, oc3, oc4, oc5 = st.columns(5)
    show_target   = oc1.checkbox("Stack Difficulty" if IS_EN else "스택 난이도", True)
    show_market   = oc2.checkbox(TR("market_board"), True)
    show_our_intg = oc3.checkbox(TR("our_intg"), True)
    show_our_board= oc4.checkbox("Our Board (normalized)" if IS_EN else "우리 board (정규화)", False)
    show_our_full = oc5.checkbox("Our Board+Stack (norm)" if IS_EN else "우리 board+스택 (정규화)", False)

    st.markdown("---")

    # ── 비교 차트
    st.subheader(TR("difficulty_curve"))

    n_all  = np.arange(1, 501)
    fig = go.Figure()

    # 스택 난이도
    if show_target:
        fig.add_trace(go.Scatter(
            x=n_all, y=baseline_curve(n_all), name="기준선 baseline(N)",
            line=dict(color="#58a6ff", width=1.5, dash="dot"), opacity=0.6
        ))
        fig.add_trace(go.Scatter(
            x=n_all, y=target_curve(n_all), name="Stack Difficulty" if IS_EN else "스택 난이도",
            line=dict(color="#58a6ff", width=2), opacity=0.9
        ))

    # 공통 가중치 설정
    W_H1 = st.session_state.get("h1_weights", {
        "H1_1":8,"H1_2":12,"H1_3":10,"H1_4":8,"H1_5":10,
        "H1_6":12,"H1_7":12,"H1_8":8,"H1_9":8,"H1_10":5,
        "H1_11":5,"H1_12":6,"H1_13":4,"H1_14":4,"H1_15":4,
    })

    # 시장 board_score
    if show_market and market is not None:
        try:
            mk = market.copy()
            mk.columns = [str(c).strip() for c in mk.columns]
            mk = mk[mk["Stage"].apply(lambda x: str(x).strip().lstrip("-").isdigit())].copy()
            mk["Stage"] = mk["Stage"].astype(int)
            tw = sum(W_H1.values())
            score = pd.Series(0.0, index=mk.index)
            for key, w in W_H1.items():
                col = MK_COL_MAP.get(key)
                if col and col in mk.columns:
                    v = pd.to_numeric(mk[col], errors="coerce").fillna(0)
                    rng = v.max() - v.min()
                    vn = (v - v.min()) / rng if rng > 0 else pd.Series(0.0, index=mk.index)
                    score += (1 - vn if W_DIR.get(key, False) else vn) * w
            mk["market_board"] = (score / tw * 100).round(1)
            mk_sm = mk["market_board"].rolling(5, center=True, min_periods=1).mean()
            fig.add_trace(go.Scatter(
                x=mk["Stage"], y=mk["market_board"].tolist(),
                name="시장 board_score",
                mode="markers+lines", marker=dict(size=4, color="#f78166"),
                line=dict(color="#f78166", width=1), opacity=0.5
            ))
            fig.add_trace(go.Scatter(
                x=mk["Stage"], y=mk_sm.tolist(),
                name="Market board_score (MA)" if IS_EN else "시장 board_score (이동평균)",
                line=dict(color="#f78166", width=2.5)
            ))
        except Exception as e:
            st.warning(f"시장 데이터 파싱 오류: {e}")

    if intg is not None:
        x_lv = list(range(1, len(intg)+1))

        # 우리 게임 통합 난이도
        if show_our_intg:
            custom    = (intg["board_score"]*w_b + intg["gameplay_score"]*w_g) / 100
            custom_sm = custom.rolling(5, center=True, min_periods=1).mean()
            fig.add_trace(go.Scatter(
                x=x_lv, y=custom.tolist(),
                name="우리 통합 (원시)", line=dict(color="#3fb950", width=1), opacity=0.4
            ))
            fig.add_trace(go.Scatter(
                x=x_lv, y=custom_sm.tolist(),
                name="Our Integrated (MA)" if IS_EN else "우리 통합 (이동평균)", line=dict(color="#3fb950", width=3)
            ))

        # 우리 게임 board_score 정규화
        if show_our_board:
            bs = intg["board_score"]
            bs_norm = ((bs - bs.min()) / (bs.max() - bs.min()) * 100) if bs.max() > bs.min() else bs
            bs_sm   = bs_norm.rolling(5, center=True, min_periods=1).mean()
            fig.add_trace(go.Scatter(
                x=x_lv, y=bs_norm.tolist(),
                name="우리 board (정규화, 원시)", line=dict(color="#d2a8ff", width=1), opacity=0.4
            ))
            fig.add_trace(go.Scatter(
                x=x_lv, y=bs_sm.tolist(),
                name="Our Board (normalized, MA)" if IS_EN else "우리 board (정규화, 이동평균)", line=dict(color="#d2a8ff", width=2.5)
            ))

        # 우리 게임 board+gameplay 정규화 (통합 정규화)
        if show_our_full:
            full = (intg["board_score"]*w_b + intg["gameplay_score"]*w_g) / 100
            full_norm = ((full - full.min()) / (full.max() - full.min()) * 100) if full.max() > full.min() else full
            full_sm   = full_norm.rolling(5, center=True, min_periods=1).mean()
            fig.add_trace(go.Scatter(
                x=x_lv, y=full_norm.tolist(),
                name="우리 board+스택 (정규화, 원시)", line=dict(color="#ffa657", width=1), opacity=0.4
            ))
            fig.add_trace(go.Scatter(
                x=x_lv, y=full_sm.tolist(),
                name="Our Board+Stack (normalized, MA)" if IS_EN else "우리 board+스택 (정규화, 이동평균)", line=dict(color="#ffa657", width=2.5)
            ))

    fig.update_layout(
        height=440, plot_bgcolor=T["plot_bg"], paper_bgcolor=T["plot_bg"],
        font_color=T["text"], xaxis_title="Level" if IS_EN else "Level" if IS_EN else "레벨", yaxis_title="난이도 점수",
        yaxis=dict(range=[0,105], gridcolor=T["grid_line"]),
        xaxis=dict(gridcolor=T["grid_line"]),
        legend=dict(orientation="h", y=1.12, bgcolor="rgba(0,0,0,0)"),
        margin=dict(l=10,r=10,t=50,b=10)
    )
    st.plotly_chart(fig, use_container_width=True)

    if market is None:
        st.info("💡 사이드바에서 시장 데이터 CSV를 업로드하면 실측값이 차트에 표시됩니다.")
    if intg is None:
        st.info("💡 사이드바에서 integrated_difficulty.csv를 업로드하면 우리 게임 곡선이 표시됩니다.")

    # ── 다운로드 (가중치 적용 결과)
    if intg is not None:
        custom    = (intg["board_score"]*w_b + intg["gameplay_score"]*w_g) / 100
        custom_sm = custom.rolling(5, center=True, min_periods=1).mean()
        result_df = intg[["board_score","gameplay_score"]].copy()
        result_df["custom_integrated"] = custom.round(2)
        result_df["custom_smoothed"]   = custom_sm.round(2)
        result_df.insert(0, "level", range(1, len(result_df)+1))
        dc1, dc2 = st.columns(2)
        dc1.download_button("📥 CSV Download" if IS_EN else "📥 CSV 다운로드", df_to_csv_bytes(result_df),
            f"integrated_w{w_b}_{w_g}.csv", "text/csv", use_container_width=True)
        dc2.download_button("📥 JSON Download" if IS_EN else "📥 JSON 다운로드", df_to_json_bytes(result_df),
            f"integrated_w{w_b}_{w_g}.json", "application/json", use_container_width=True)

    st.markdown("---")

    # ── 하단: 우리 난이도 구간 분석
    st.subheader(TR("zone_analysis"))
    if True:
        if intg is None:
            st.warning("Please upload integrated_difficulty.csv from the sidebar." if IS_EN else "사이드바에서 integrated_difficulty.csv를 업로드해주세요.")
        else:
            # 현재 가중치로 통합 난이도 재계산
            w_b = st.session_state.w_board
            w_g = st.session_state.w_gameplay
            custom_intg = (intg["board_score"] * w_b + intg["gameplay_score"] * w_g) / 100

            c1, c2, c3 = st.columns(3)
            c1.markdown(f'<div class="metric-card"><div class="metric-val">{custom_intg.mean():.1f}</div><div class="metric-lbl">{"Avg Integrated" if IS_EN else "평균 통합 난이도"}</div></div>', unsafe_allow_html=True)
            c2.markdown(f'<div class="metric-card"><div class="metric-val">{custom_intg.max():.1f}</div><div class="metric-lbl">{"Peak" if IS_EN else "최고점"} (Lv{custom_intg.idxmax()+1})</div></div>', unsafe_allow_html=True)
            c3.markdown(f'<div class="metric-card"><div class="metric-val">{custom_intg.min():.1f}</div><div class="metric-lbl">{"Bottom" if IS_EN else "최저점"} (Lv{custom_intg.idxmin()+1})</div></div>', unsafe_allow_html=True)
            st.markdown("")

            zone_size = st.select_slider("Zone Size" if IS_EN else "구간 크기", [10, 25, 50, 100], value=50, key="zone_sz")

            # 구간별 집계
            zones = []
            for i in range(0, len(intg), zone_size):
                s2   = intg.iloc[i:i+zone_size]
                ci   = custom_intg.iloc[i:i+zone_size]
                label = f"Lv{i+1}-{min(i+zone_size, len(intg))}"
                # 누적 막대용: 판 모양 기여 / 게임 진행 기여 (가중치 적용)
                board_contrib    = round(s2["board_score"].mean() * w_b / 100, 1)
                gameplay_contrib = round(s2["gameplay_score"].mean() * w_g / 100, 1)
                integrated_avg   = round(ci.mean(), 1)
                zones.append({
                    "Zone" if IS_EN else "구간": label,
                    "Board Contribution" if IS_EN else "Board Contribution" if IS_EN else "판 모양 기여": board_contrib,
                    "Gameplay Contribution" if IS_EN else "Gameplay Contribution" if IS_EN else "게임 진행 기여": gameplay_contrib,
                    "Integrated Avg" if IS_EN else "통합 평균": integrated_avg,
                    "Board (raw)": round(s2["board_score"].mean(), 1),
                    "Gameplay (raw)": round(s2["gameplay_score"].mean(), 1),
                })
            zdf = pd.DataFrame(zones)

            # 누적 막대그래프 + target(N) 기준선
            import math, numpy as np
            LOCAL_VAR = [-4.18,-10.77,-9.35,0.77,-12.51,-24.5,16.84,7.51,-23.07,-4.67,49.59,28.79,-0.01,-7.79,-21.15,0.89,26.65,31.29,18.55,10.54,64.81,44.5,-0.6,26.22,18.69,15.97,18.07,48.05,15.12,46.61,-0.1,36.66,9.08,-1.83,-10.36,-9.82,16.77,-5.48,-2.13,0.26,-35.93,-12.55,-8.93,8.39,23.32,-11.53,3.89,-13.47,8.79,14.44,-19.23,-9.61,-6.62,-0.31,-15.68,-45.8,-17.16,8.99,-15.73,6.0,-12.44,-5.44,33.37,7.2,2.44,-12.58,-8.38,22.13,-14.12,-6.12,-17.77,-24.71,32.97,-11.54,1.42,-8.55,-1.06,-9.44,7.65,-0.38,-34.71,-29.79,-21.23,-26.16,2.33,12.95,-16.47,-34.48,5.62,9.64,-15.79,-14.75,44.79,5.4,-39.42,-16.52,-20.26,-44.14,1.36,-8.16]
            def tgt(N): return float(np.clip(70-52*math.exp(-N/90)+3.71+LOCAL_VAR[(N-1)%100],0,100))

            # 구간별 target 평균
            target_avgs = []
            for i in range(0, len(intg), zone_size):
                lo = i+1; hi = min(i+zone_size, len(intg))
                target_avgs.append(round(sum(tgt(n) for n in range(lo, hi+1))/(hi-lo+1), 1))
            zdf["Stack Difficulty" if IS_EN else "스택 난이도"] = target_avgs

            fig3 = go.Figure()

            # 누적 막대: 판 모양 기여 (아래)
            fig3.add_trace(go.Bar(
                x=zdf["Zone" if IS_EN else "구간"], y=zdf["Board Contribution" if IS_EN else "Board Contribution" if IS_EN else "판 모양 기여"],
                name=f"판 모양 ({w_b}%)",
                marker_color="#fa8c16",
                text=zdf["Board Contribution" if IS_EN else "Board Contribution" if IS_EN else "판 모양 기여"].apply(lambda v: f"{v:.1f}"),
                textposition="inside",
                textfont=dict(size=10, color="white"),
            ))
            # 누적 막대: 게임 진행 기여 (위)
            fig3.add_trace(go.Bar(
                x=zdf["Zone" if IS_EN else "구간"], y=zdf["Gameplay Contribution" if IS_EN else "Gameplay Contribution" if IS_EN else "게임 진행 기여"],
                name=f"게임 진행 ({w_g}%)",
                marker_color="#1890ff",
                text=zdf["Gameplay Contribution" if IS_EN else "Gameplay Contribution" if IS_EN else "게임 진행 기여"].apply(lambda v: f"{v:.1f}"),
                textposition="inside",
                textfont=dict(size=10, color="white"),
            ))
            # 통합 평균 꺾은선
            fig3.add_trace(go.Scatter(
                x=zdf["Zone" if IS_EN else "구간"], y=zdf["Integrated Avg" if IS_EN else "통합 평균"],
                name="Integrated Difficulty" if IS_EN else "통합 난이도", mode="lines+markers+text",
                line=dict(color="#3fb950", width=2.5),
                marker=dict(size=7, color="#3fb950"),
                text=zdf["Integrated Avg" if IS_EN else "통합 평균"].apply(lambda v: f"{v:.1f}"),
                textposition="top center",
                textfont=dict(size=10, color="#3fb950"),
            ))
            # target(N) 기준선
            fig3.add_trace(go.Scatter(
                x=zdf["Zone" if IS_EN else "구간"], y=zdf["Stack Difficulty" if IS_EN else "스택 난이도"],
                name="Stack Difficulty" if IS_EN else "스택 난이도", mode="lines+markers",
                line=dict(color="#f5222d", width=1.5, dash="dot"),
                marker=dict(size=5, color="#f5222d"),
            ))

            fig3.update_layout(
                height=380, barmode="stack",
                plot_bgcolor=T["plot_bg"], paper_bgcolor=T["plot_bg"],
                font_color=T["text"],
                yaxis=dict(range=[0, 105], gridcolor=T["grid_line"], title="Difficulty" if IS_EN else "난이도"),
                xaxis=dict(gridcolor=T["grid_line"]),
                legend=dict(orientation="h", y=1.12, bgcolor="rgba(0,0,0,0)"),
                margin=dict(l=10, r=10, t=40, b=10),
            )
            st.plotly_chart(fig3, use_container_width=True)

            # 표 (raw 값 포함)
            show_cols = ["Zone" if IS_EN else "구간", "Board (raw)", "Gameplay (raw)", "Integrated Avg" if IS_EN else "통합 평균", "Stack Difficulty" if IS_EN else "스택 난이도"]
            st.dataframe(zdf[show_cols].rename(columns={
                "Board (raw)": f"판 모양 ({w_b}%)",
                "Gameplay (raw)": f"게임 진행 ({w_g}%)",
            }), use_container_width=True)


# ══════════════════════════════════════════════════════
elif page == "🗺️ 3. 판 모양 뷰어":
    st.markdown('<div class="page-anim">', unsafe_allow_html=True)
    st.title(TR("viewer_title"))

    view_tab, edit_tab = st.tabs([TR("lv_viewer"), TR("new_board")])

    # ── 뷰어 & 인라인 편집
    with view_tab:
        # ── 소스 선택
        vc1, vc2 = st.columns([1,3])
        with vc1:
            src = st.radio("Source" if IS_EN else TR("source"), [TR("lv_number"),TR("json_upload")], horizontal=True)
            data = None
            fname_default = "N_001.json"
            if src == TR("lv_number"):
                lv = st.number_input("Level" if IS_EN else "Level" if IS_EN else "레벨", 1, 500, 1)
                fname_default = f"N_{int(lv):03d}.json"
                data = load_level_local(int(lv))
                if data is None:
                    st.warning(f"N_{lv:03d}.json 없음 — JSON 업로드 사용")
            else:
                uj = st.file_uploader("JSON 파일", type=["json"], key="uj_view")
                if uj:
                    data = json.load(uj)
                    fname_default = uj.name
                    st.success(f"✅ {uj.name}")

            # 편집 모드 토글
            edit_mode = st.toggle("✏️ 편집 모드", value=False, key="view_edit_mode")

            if data:
                data_key      = f"view_tiles_{fname_default}"
                data_orig_key = f"view_tiles_orig_{fname_default}"

                # 최초 로드 또는 편집 모드 OFF → 원본 보존 + 작업본 초기화
                if data_orig_key not in st.session_state:
                    st.session_state[data_orig_key] = json.loads(json.dumps(data))
                if data_key not in st.session_state or not edit_mode:
                    st.session_state[data_key] = json.loads(json.dumps(data))

                # 초기값 복원 버튼 (편집 모드일 때만)
                if edit_mode:
                    if st.button("🔄 초기값으로 복원", use_container_width=True, key="ve_reset"):
                        st.session_state[data_key] = json.loads(
                            json.dumps(st.session_state[data_orig_key])
                        )
                        st.success("초기값으로 복원됐어요!")
                        st.rerun()

                h1 = analyze_level(data)
                st.markdown(f"**{'Board' if IS_EN else '보드'}**: {data['XCells']}×{data['YCells']}")
                with st.expander("Tile Composition" if IS_EN else "Tile Composition" if IS_EN else "타일 구성"):
                    for k,v in h1['tile_counts'].items():
                        if v>0 and k!='Blank':
                            st.markdown(f"- {k}: {v}개")
                with st.expander("H1 Indicators" if IS_EN else "H1 Indicators" if IS_EN else "H1 지표"):
                    for k in ['H1_1','H1_2','H1_3','H1_5','H1_6','H1_7','H1_9','H1_12','H1_14']:
                        st.markdown(f"**{k}**: {h1[k]}")
                h1e = {k:v for k,v in h1.items() if k!='tile_counts'}
                h1df = pd.DataFrame([h1e])
                st.download_button("📥 H1 CSV", df_to_csv_bytes(h1df),
                                   "h1_metrics.csv", "text/csv", use_container_width=True)

            show_coord = st.checkbox("Show Coordinates" if IS_EN else "Show Coordinates" if IS_EN else TR("show_coord"), False)
            show_chips = st.checkbox("Show Chip Colors" if IS_EN else TR("show_chip"), True)
            hex_size   = st.slider("헥사 크기", 20, 60, 38)

        with vc2:
            if data is None:
                st.info("왼쪽에서 레벨을 선택하거나 JSON을 업로드해주세요.")
            else:
                work_data = st.session_state.get(f"view_tiles_{fname_default}", data)
                Y = work_data['YCells']; X = work_data['XCells']
                tiles = work_data['Tiles']

                # ── 편집 모드: 셀 선택 패널
                if edit_mode:
                    ep1, ep2 = st.columns(2)
                    e_sel_y = ep1.number_input("행(Y)", 0, Y-1, 0, key="ve_sel_y")
                    e_sel_x = ep2.number_input("열(X)", 0, X-1, 0, key="ve_sel_x")
                    cur = tiles[e_sel_y][e_sel_x]
                    cur_type = TILETYPE.get(cur.get('TileType',0), 'Normal')

                    ep3, ep4 = st.columns(2)
                    new_type = ep3.selectbox("TileType", list(TILETYPE.values()),
                                             index=list(TILETYPE.values()).index(cur_type),
                                             key="ve_type")
                    new_stacks, lv_val, ul_val = [], 0, 0
                    if new_type in ('Stack','StackLock','Ice'):
                        st.markdown(
                            "".join([f'<span style="background:{CHIP_HEX[i]};color:{"#333" if i in(1,6) else "white"};border-radius:3px;padding:1px 5px;margin:1px;font-size:10px;">{i}={COLOR_MAP[i][:3]}</span>' for i in range(8)]),
                            unsafe_allow_html=True)
                        si = ep4.text_input("칩 (콤마구분)",
                                            ",".join(map(str, cur.get('Stacks',[]))), key="ve_stacks")
                        try: new_stacks=[int(s) for s in si.split(',') if s.strip().isdigit()]
                        except: new_stacks=[]
                    if new_type in ('Lock','Plank'):
                        lv_val = ep4.number_input("Level", 0, 9999, cur.get('Level',0), key="ve_lv")
                    if new_type in ('StackLock','Ice'):
                        ul_val = ep4.number_input("UnlockLevel", 0, 9999, cur.get('UnlockLevel',0), key="ve_ul")

                    if st.button("✅ Apply" if IS_EN else TR("apply_cell"), key="ve_apply"):
                        new_cell = {'TileType': TILE_REV[new_type]}
                        if new_type in ('Stack','StackLock','Ice'): new_cell['Stacks'] = new_stacks
                        if new_type in ('Lock','Plank'): new_cell['Level'] = lv_val
                        if new_type in ('StackLock','Ice'): new_cell['UnlockLevel'] = ul_val
                        st.session_state[f"view_tiles_{fname_default}"]['Tiles'][e_sel_y][e_sel_x] = new_cell
                        st.rerun()

                # ── 그리드 렌더링
                fig = go.Figure()
                for y in range(Y):
                    for x in range(X):
                        tile = tiles[y][x]; tt = tile.get('TileType',0)
                        name = TILETYPE.get(tt,'Normal')
                        cx, cy = hex_to_pixel(y, x, hex_size)
                        hx, hy = make_hex_path(cx, cy, hex_size*0.92)

                        is_sel = edit_mode and (y==e_sel_y and x==e_sel_x) if edit_mode else False

                        if name == 'Blank' and not is_sel:
                            fig.add_trace(go.Scatter(x=hx,y=hy,fill='toself',
                                fillcolor='rgba(0,0,0,0)',
                                line=dict(color='rgba(0,0,0,0)',width=0),
                                mode='lines',hoverinfo='skip',showlegend=False))
                            continue

                        border_c = '#FFD700' if is_sel else 'white'
                        border_w = 3 if is_sel else 1.5
                        fill_c   = HEX_COLORS.get(name,'#CCC') if name!='Blank' else 'rgba(80,80,80,0.3)'

                        fig.add_trace(go.Scatter(x=hx,y=hy,fill='toself',
                            fillcolor=fill_c,
                            line=dict(color=border_c,width=border_w),
                            mode='lines',hoverinfo='skip',showlegend=False))

                        label = name[:2]
                        if name in ('Stack','StackLock','Ice') and 'Stacks' in tile:
                            stacks = tile['Stacks']
                            label = '+'.join(COLOR_MAP.get(c,'?')[0] for c in stacks[:4]) if show_chips and stacks else f"S{len(stacks)}"
                        elif name in ('Lock','Plank') and 'Level' in tile:
                            label = f"L{tile['Level']}"
                        elif name == 'StackLock' and 'UnlockLevel' in tile:
                            label = f"SL{tile['UnlockLevel']}"
                        if show_coord or edit_mode:
                            label = f"({y},{x})\n{label}"

                        fig.add_annotation(x=cx,y=cy,text=label,showarrow=False,
                            font=dict(size=9,color='white' if tt!=0 else '#333'),align='center')

                fig.update_layout(height=600,margin=dict(l=10,r=10,t=10,b=10),
                    xaxis=dict(visible=False,scaleanchor='y'),
                    yaxis=dict(visible=False),
                    plot_bgcolor=T["grid_bg"],paper_bgcolor=T["grid_bg"],showlegend=False)
                st.plotly_chart(fig, use_container_width=True)

                # ── 저장 버튼 (편집 모드일 때만)
                if edit_mode:
                    st.markdown("---")
                    save_fname = st.text_input("저장 파일명", fname_default, key="ve_fname")
                    json_out = json.dumps(work_data, ensure_ascii=False, indent=2).encode("utf-8")

                    sc1, sc2 = st.columns(2)
                    sc1.download_button("📥 JSON Download" if IS_EN else "📥 JSON 다운로드", json_out,
                    save_fname, "application/json", use_container_width=True,
                    key="dl_json_view")

                    if sc2.button("☁️ GitHub에 저장", use_container_width=True, key="ve_github"):
                        token = st.session_state.github_token
                        if not token:
                            st.error("사이드바 → 🔑 GitHub 설정에서 토큰을 입력해주세요.")
                        else:
                            gh_path = f"data/levels/{save_fname}"
                            with st.spinner(f"GitHub에 저장 중... ({gh_path})"):
                                ok, resp = github_commit(
                                    token, GITHUB_REPO, gh_path, json_out,
                                    f"update level: {save_fname}"
                                )
                            if ok:
                                st.success(f"✅ GitHub 저장 완료! `{gh_path}`")
                                # 캐시 초기화
                                load_level_local.clear()
                            else:
                                st.error(f"❌ 실패: {resp.get('message','')}")

    # ── 편집기
    with edit_tab:
        st.caption("Set XCells·YCells → Select cell → Set type → Apply → Save JSON" if IS_EN else "XCells·YCells 설정 → 행/열로 셀 선택 → 타입 지정 → 셀 적용 → JSON 저장")

        # 그리드 크기 설정
        ec1,ec2,ec3 = st.columns([1,1,2])
        new_x = ec1.number_input("XCells",3,10,st.session_state.grid_x,key="nx")
        new_y = ec2.number_input("YCells",3,10,st.session_state.grid_y,key="ny")
        if ec3.button("🔄 Reset Grid" if IS_EN else TR("reset_grid")) or st.session_state.grid_tiles is None \
                or len(st.session_state.grid_tiles)!=new_y \
                or len(st.session_state.grid_tiles[0])!=new_x:
            st.session_state.grid_x = new_x
            st.session_state.grid_y = new_y
            st.session_state.grid_tiles = [[{'TileType':0} for _ in range(new_x)] for _ in range(new_y)]

        tiles_e = st.session_state.grid_tiles
        Y_e, X_e = new_y, new_x

        # ── 메인 레이아웃: 그리드(좌) + 정보패널(우)
        gcol, icol = st.columns([3, 1])

        with icol:
            # ── 셀 선택
            st.markdown("### 📍 Cell Selection" if IS_EN else "### 📍 셀 선택")
            sel_y = st.number_input("Row (Y, top→bottom)" if IS_EN else TR("row_sel"), 0, Y_e-1, 0, key="sel_y")
            sel_x = st.number_input("Column (X, left→right)" if IS_EN else TR("col_sel"),  0, X_e-1, 0, key="sel_x")

            cur_tile = tiles_e[sel_y][sel_x]
            cur_type = TILETYPE.get(cur_tile.get('TileType', 0), 'Normal')

            st.markdown("---")

            # ── 현재 셀 정보 표시
            st.markdown("### 🔍 Current Cell Info" if IS_EN else "### 🔍 현재 셀 정보")
            tile_color = HEX_COLORS.get(cur_type, '#CCC')
            st.markdown(
                f'<div style="background:{tile_color};border-radius:8px;'
                f'padding:8px 12px;margin-bottom:8px;font-weight:700;'
                f'color:{"#333" if cur_type=="Normal" else "white"};font-size:14px;">'
                f'({sel_y}, {sel_x}) — {cur_type}</div>',
                unsafe_allow_html=True
            )

            # 칩 색상 미리보기
            if cur_type in ('Stack', 'StackLock', 'Ice') and 'Stacks' in cur_tile:
                stacks_now = cur_tile['Stacks']
                st.markdown(f"**칩 수**: {len(stacks_now)}개")
                chip_html = ""
                for c in stacks_now:
                    cname = COLOR_MAP.get(c, '?')
                    chex  = CHIP_HEX.get(c, '#888')
                    chip_html += (
                        f'<span style="display:inline-block;background:{chex};'
                        f'color:{"#333" if c in (1,6) else "white"};'
                        f'border-radius:4px;padding:2px 7px;margin:2px;'
                        f'font-size:11px;font-weight:600;">{c} {cname}</span>'
                    )
                st.markdown(chip_html, unsafe_allow_html=True)
            elif cur_type in ('Lock', 'Plank'):
                st.markdown(f"**Level**: {cur_tile.get('Level', 0)}")
            elif cur_type in ('StackLock', 'Ice'):
                st.markdown(f"**UnlockLevel**: {cur_tile.get('UnlockLevel', 0)}")
                if 'Stacks' in cur_tile:
                    st.markdown(f"**칩 수**: {len(cur_tile['Stacks'])}개")

            # 원시 JSON
            with st.expander("raw JSON"):
                st.markdown(
                    f'<pre style="background:#1a1a2e;color:#FFFFFF;padding:12px;'
                    f'border-radius:6px;font-size:12px;overflow-x:auto;">'
                    f'{json.dumps(cur_tile, ensure_ascii=False, indent=2)}'
                    f'</pre>',
                    unsafe_allow_html=True
                )

            st.markdown("---")

            # ── 셀 편집
            st.markdown("### ✏️ Cell Edit" if IS_EN else "### ✏️ 셀 편집")
            new_type = st.selectbox("TileType", [
                'Normal','Blank','Stack','Lock','Plank',
                'Ice','StackLock','Grass','Ads','CameraPicture'
            ], index=list(TILETYPE.values()).index(cur_type), key="sel_type")

            new_stacks, lv_val, ul_val = [], 0, 0

            if new_type in ('Stack', 'StackLock', 'Ice'):
                st.markdown("**Chip Colors** (0~7, comma-separated)" if IS_EN else "**칩 색상** (0~7, 콤마구분)")
                # 색상 참조표
                st.markdown(
                    "".join([
                        f'<span style="background:{CHIP_HEX[i]};color:{"#333" if i in (1,6) else "white"};'
                        f'border-radius:3px;padding:1px 5px;margin:1px;font-size:10px;">'
                        f'{i}={COLOR_MAP[i][:3]}</span>'
                        for i in range(8)
                    ]),
                    unsafe_allow_html=True
                )
                stacks_input = st.text_input(
                    "칩 입력",
                    ",".join(map(str, cur_tile.get('Stacks', []))),
                    key="sel_stacks",
                    help="예: 0,1,2,0  → Blue, Yellow, Red, Blue 순서로 쌓임"
                )
                try:
                    new_stacks = [int(s) for s in stacks_input.split(',') if s.strip().isdigit()]
                except:
                    new_stacks = []
                # 입력 미리보기
                if new_stacks:
                    preview_html = "".join([
                        f'<span style="background:{CHIP_HEX.get(c,"#888")};'
                        f'color:{"#333" if c in (1,6) else "white"};'
                        f'border-radius:3px;padding:2px 6px;margin:2px;font-size:11px;">'
                        f'{COLOR_MAP.get(c,"?")}</span>'
                        for c in new_stacks
                    ])
                    st.markdown(preview_html, unsafe_allow_html=True)

            if new_type in ('Lock', 'Plank'):
                lv_val = st.number_input("Level", 0, 9999,
                                         cur_tile.get('Level', 0), key="sel_lv")
            if new_type in ('StackLock', 'Ice'):
                ul_val = st.number_input("UnlockLevel", 0, 9999,
                                          cur_tile.get('UnlockLevel', 0), key="sel_ul")

            if st.button("✅ Apply" if IS_EN else TR("apply_cell"), use_container_width=True):
                new_cell = {'TileType': TILE_REV[new_type]}
                if new_type in ('Stack', 'StackLock', 'Ice'):
                    new_cell['Stacks'] = new_stacks
                if new_type in ('Lock', 'Plank'):
                    new_cell['Level'] = lv_val
                if new_type in ('StackLock', 'Ice'):
                    new_cell['UnlockLevel'] = ul_val
                st.session_state.grid_tiles[sel_y][sel_x] = new_cell
                st.rerun()

            st.markdown("---")
            # ── JSON 저장
            st.markdown("### 💾 Save" if IS_EN else "### 💾 저장")
            fname = st.text_input("파일명", "N_001.json", key="edit_fname")
            json_out = json.dumps({
                "Timestamp": int(datetime.now().timestamp()*1000),
                "GameType": 0, "GridOrientation": 0,
                "XCells": X_e, "YCells": Y_e,
                "Tiles": st.session_state.grid_tiles
            }, ensure_ascii=False, indent=2).encode("utf-8")
            st.download_button("📥 JSON Download" if IS_EN else "📥 JSON 다운로드", json_out, fname,
                   "application/json", use_container_width=True,
                   key="dl_json_edit")

        # ── 그리드 미리보기 (좌측)
        with gcol:
            fig_e = go.Figure()
            hs = 42
            for y in range(Y_e):
                for x in range(X_e):
                    tile = tiles_e[y][x]
                    tt   = tile.get('TileType', 0)
                    name = TILETYPE.get(tt, 'Normal')
                    cx, cy = hex_to_pixel(y, x, hs)
                    hx, hy = make_hex_path(cx, cy, hs * 0.92)

                    is_sel = (y == sel_y and x == sel_x)

                    if name == 'Blank' and not is_sel:
                        # 투명하게 자리만 유지 (선택된 Blank는 금색으로 표시)
                        fig_e.add_trace(go.Scatter(
                            x=hx, y=hy, fill='toself',
                            fillcolor='rgba(0,0,0,0)',
                            line=dict(color='rgba(0,0,0,0)', width=0),
                            mode='lines', hoverinfo='skip', showlegend=False
                        ))
                        continue

                    fill         = HEX_COLORS.get(name, '#CCC') if name != 'Blank' else 'rgba(80,80,80,0.3)'
                    border_color = '#FFD700' if is_sel else 'white'
                    border_w     = 4 if is_sel else 1.5

                    fig_e.add_trace(go.Scatter(
                        x=hx, y=hy, fill='toself',
                        fillcolor=fill,
                        line=dict(color=border_color, width=border_w),
                        mode='lines', hoverinfo='skip', showlegend=False
                    ))

                    # 셀 라벨
                    if name == 'Blank':
                        label = 'Bl'
                    elif name in ('Stack', 'StackLock', 'Ice') and 'Stacks' in tile:
                        stacks = tile['Stacks']
                        # 칩 색 이니셜 표시
                        label = '+'.join(COLOR_MAP.get(c,'?')[0] for c in stacks[:3])
                        if len(stacks) > 3:
                            label += f'+{len(stacks)-3}↑'
                    elif name in ('Lock', 'Plank') and 'Level' in tile:
                        label = f"L{tile['Level']}"
                    elif name == 'StackLock' and 'UnlockLevel' in tile:
                        label = f"SL\n{tile['UnlockLevel']}"
                    else:
                        label = name[:2]

                    # 좌표 항상 표시 (편집기에선 유용)
                    coord_label = f"({y},{x})\n{label}"
                    font_color  = 'white' if tt not in (0,) else '#444'

                    fig_e.add_annotation(
                        x=cx, y=cy, text=coord_label,
                        showarrow=False,
                        font=dict(size=9, color=font_color),
                        align='center'
                    )

            fig_e.update_layout(
                height=560,
                margin=dict(l=10, r=10, t=10, b=10),
                xaxis=dict(visible=False, scaleanchor='y'),
                yaxis=dict(visible=False),
                plot_bgcolor=T["grid_bg"],
                paper_bgcolor=T["grid_bg"],
                showlegend=False
            )
            st.plotly_chart(fig_e, use_container_width=True)

            # ── 실시간 난이도 계산
            level_data_now = {
                'XCells': X_e, 'YCells': Y_e,
                'Tiles': st.session_state.grid_tiles
            }
            h1_now = analyze_level(level_data_now)

            _default_h1 = {
                'H1_1':8,'H1_2':12,'H1_3':10,'H1_4':8,'H1_5':10,
                'H1_6':12,'H1_7':12,'H1_8':8,'H1_9':8,'H1_10':5,
                'H1_11':5,'H1_12':6,'H1_13':4,'H1_14':4,'H1_15':4
            }
            _h1_base = st.session_state.get('h1_weights', _default_h1)
            W_H1_now = {
                key: st.session_state.get(f"w_{key}", _h1_base.get(key, 8))
                for key in _default_h1
            }
            W_DIR_now = {
                "H1_1":True,"H1_2":True,"H1_3":True,"H1_4":True,"H1_5":False,
                "H1_6":False,"H1_7":False,"H1_8":False,"H1_9":False,"H1_10":False,
                "H1_11":False,"H1_12":False,"H1_13":True,"H1_14":True,"H1_15":True,
            }
            tw_now = sum(W_H1_now.values())
            score_now = 0.0

            # 시장 데이터 기준 정규화를 위한 참조값 (근사치)
            H1_REF_MIN = {"H1_1":4,"H1_2":0,"H1_3":0,"H1_4":0,"H1_5":0,
                          "H1_6":0,"H1_7":0,"H1_8":0,"H1_9":0,"H1_10":0,
                          "H1_11":0,"H1_12":0,"H1_13":0,"H1_14":0,"H1_15":0}
            H1_REF_MAX = {"H1_1":64,"H1_2":200,"H1_3":30,"H1_4":80,"H1_5":15,
                          "H1_6":80,"H1_7":60,"H1_8":60,"H1_9":20,"H1_10":40,
                          "H1_11":10,"H1_12":5000,"H1_13":20,"H1_14":3,"H1_15":40}

            for k, w in W_H1_now.items():
                v   = h1_now.get(k, 0)
                lo  = H1_REF_MIN.get(k, 0)
                hi  = H1_REF_MAX.get(k, 1)
                rng_v = hi - lo if hi > lo else 1
                vn  = max(0.0, min(1.0, (v - lo) / rng_v))
                if W_DIR_now.get(k, False):
                    vn = 1 - vn
                score_now += vn * w

            board_score_now = round(score_now / tw_now * 100, 1)
            grade_now = ('매우쉬움' if board_score_now < 25 else
                         '쉬움'     if board_score_now < 45 else
                         '보통'     if board_score_now < 60 else
                         '어려움'   if board_score_now < 75 else '매우어려움')
            grade_color = {'매우쉬움':'#1890FF','쉬움':'#52C41A','보통':'#FADB14',
                           '어려움':'#FA8C16','매우어려움':'#F5222D'}[grade_now]
            grade_emoji = {'매우쉬움':'🔵','쉬움':'🟢','보통':'🟡',
                           '어려움':'🟠','매우어려움':'🔴'}[grade_now]

            st.markdown(
                f"""<div style="background:rgba(255,255,255,0.05);border:1px solid rgba(255,255,255,0.1);
                border-left:4px solid {grade_color};border-radius:10px;
                padding:12px 16px;margin:12px 0;">
                <div style="font-size:12px;color:#9ca3af;margin-bottom:4px;">📐 현재 판 모양 난이도</div>
                <div style="display:flex;align-items:center;gap:12px;">
                  <span style="font-size:28px;font-weight:700;color:{grade_color};">{board_score_now}</span>
                  <span style="font-size:18px;">{grade_emoji}</span>
                  <span style="font-size:16px;font-weight:600;color:{grade_color};">{grade_now}</span>
                </div>
                </div>""",
                unsafe_allow_html=True
            )

            # 타일 범례
            st.markdown("**타일 범례**")
            legend_html = "".join([
                f'<span style="background:{HEX_COLORS[t]};color:{"#333" if t=="Normal" else "white"};'
                f'border-radius:4px;padding:2px 8px;margin:3px;font-size:11px;display:inline-block;">'
                f'{t}</span>'
                for t in HEX_COLORS if t != 'Blank'
            ])
            st.markdown(legend_html, unsafe_allow_html=True)

# ══════════════════════════════════════════════════════
# 탭 4 — 설정

# ══════════════════════════════════════════════════════
# 탭 4 — JSON 생성기
# ══════════════════════════════════════════════════════
elif page == "🎲 4. JSON 생성기":
    st.markdown('<div class="page-anim">', unsafe_allow_html=True)
    st.title(TR("gen_title"))
    st.caption(TR("gen_cap"))

    from generate_levels import generate_range_zip, target_diff as calc_diff

    tbl = st.session_state.tbl_df

    if tbl is None:
        st.warning("Please upload tblStage_500.xlsx from the sidebar." if IS_EN else "사이드바에서 tblStage_500.xlsx를 업로드해주세요.")
    else:
        # ── 난이도 곡선 미리보기
        st.subheader(TR("gen_curve"))
        lv_range = st.slider(TR("lv_range"), 1, 500, (1, 50), key="gen_range")
        start_lv, end_lv = lv_range
        total_lv = end_lv - start_lv + 1

        # 선택 구간 난이도 계산
        all_lvs   = list(range(1, 501))
        all_diffs = [calc_diff(n) for n in all_lvs]

        fig_gen = go.Figure()
        # 전체 곡선 (연하게)
        fig_gen.add_trace(go.Scatter(
            x=all_lvs, y=all_diffs,
            mode='lines', name='전체 곡선',
            line=dict(color='#444', width=1), opacity=0.4
        ))
        # 선택 구간 강조
        sel_lvs   = list(range(start_lv, end_lv + 1))
        sel_diffs = all_diffs[start_lv-1:end_lv]
        fig_gen.add_trace(go.Scatter(
            x=sel_lvs, y=sel_diffs,
            mode='lines+markers', name=f'선택 구간 (Lv {start_lv}~{end_lv})',
            line=dict(color='#3fb950', width=2.5),
            marker=dict(size=4, color='#3fb950')
        ))
        fig_gen.update_layout(
            height=320,
            plot_bgcolor=T["plot_bg"], paper_bgcolor=T["plot_bg"],
            font_color=T["text"],
            xaxis=dict(title="Level" if IS_EN else "Level" if IS_EN else "레벨", gridcolor=T["grid_line"]),
            yaxis=dict(title="Difficulty" if IS_EN else "난이도", range=[0,105], gridcolor=T["grid_line"]),
            legend=dict(orientation="h", y=1.1, bgcolor="rgba(0,0,0,0)"),
            margin=dict(l=10,r=10,t=30,b=10)
        )
        st.plotly_chart(fig_gen, use_container_width=True)

        # ── 선택 구간 요약
        sel_arr = sel_diffs
        c1, c2, c3, c4 = st.columns(4)
        c1.metric("Levels to Generate" if IS_EN else "Levels to Generate" if IS_EN else "생성할 레벨 수", f"{total_lv}개")
        c2.metric("Avg Difficulty" if IS_EN else TR("avg_diff"), f"{sum(sel_arr)/len(sel_arr):.1f}")
        c3.metric("Min" if IS_EN else "최저", f"{min(sel_arr):.1f} (Lv{start_lv + sel_arr.index(min(sel_arr))})")
        c4.metric("Max" if IS_EN else "최고", f"{max(sel_arr):.1f} (Lv{start_lv + sel_arr.index(max(sel_arr))})")

        st.markdown("---")

        # ── 생성 버튼
        if st.button(f"🚀 Generate Lv {start_lv}~{end_lv} JSON ({total_lv} files)" if IS_EN else f"🚀 Lv {start_lv}~{end_lv} JSON 생성 ({total_lv}개)", type="primary", use_container_width=True):

            progress_bar = st.progress(0)
            status_text  = st.empty()

            GRADE_EMOJI = {
                'very_easy': ('매우쉬움', '🔵'),
                'easy':      ('쉬움',     '🟢'),
                'normal':    ('보통',     '🟡'),
                'hard':      ('어려움',   '🟠'),
                'very_hard': ('매우어려움','🔴'),
            }
            def on_progress(done, total, lv_now=0, bs=0, ss=0, intg_v=0):
                if lv_now == 0: lv_now = start_lv + done - 1
                diff_now = calc_diff(lv_now)
                g = ('very_easy' if diff_now < 25 else
                     'easy'      if diff_now < 45 else
                     'normal'    if diff_now < 60 else
                     'hard'      if diff_now < 75 else 'very_hard')
                grade_name, emoji = GRADE_EMOJI[g]
                progress_bar.progress(done / total)
                status_text.markdown(
                    f"⚙️ **Lv {lv_now}** {"Generating" if IS_EN else "생성 중"} &nbsp;|&nbsp; "
                    f"{"Difficulty" if IS_EN else "Difficulty" if IS_EN else "난이도"} **{diff_now}** &nbsp; {emoji} {grade_name} &nbsp;|&nbsp; "
                    f"({done}/{total})"
                )

            df_n = tbl[
                tbl['LevelName'].str.startswith('N_', na=False) |
                tbl['LevelName'].str.startswith('N ', na=False)
            ].reset_index(drop=True)

            if df_n.empty:
                st.error("tblStage에서 N_ 레벨 데이터를 찾을 수 없어요. LevelName 컬럼을 확인해주세요.")
                st.stop()

            zip_bytes = generate_range_zip(start_lv, end_lv, df_n, callback=on_progress)

            progress_bar.progress(1.0)
            status_text.markdown(f"✅ **Generation Complete!**" if IS_EN else "✅ **{} files generated!**".format(total_lv) if IS_EN else "✅ **{}개 생성 완료!**".format(total_lv))

            st.download_button(
                label=f"📥 Download ZIP (N_{start_lv:03d} ~ N_{end_lv:03d}.json)" if IS_EN else f"📥 ZIP 다운로드 (N_{start_lv:03d} ~ N_{end_lv:03d}.json)",
                data=zip_bytes,
                file_name=f"levels_{start_lv:03d}_{end_lv:03d}.zip",
                mime="application/zip",
                use_container_width=True,
            )


# ══════════════════════════════════════════════════════
elif page == "🔧 5. 설정":
    st.markdown('<div class="page-anim">', unsafe_allow_html=True)
    st.title(TR("settings_title"))

    set_tab1, set_tab2, set_tab3 = st.tabs([(TR("h1_weight_tab")), (TR("tbl_weight_tab")), (TR("stack_edit_tab"))])

    # ── 가중치 세부 설정
    with set_tab1:
        st.caption(TR("h1_weight_cap"))

        H1_INFO = [
            ('H1_1','전체 그리드 수',True),
            ('H1_2','Normal 열린 변 합',True),
            ('H1_3','Normal 개수',True),
            ('H1_4','Stack 열린 변 합',True),
            ('H1_5','Stack 개수',False),
            ('H1_6','타일 색 총합',False),
            ('H1_7','색 변화 복잡도',False),
            ('H1_8','Lock 열린 변 합',False),
            ('H1_9','Lock 개수',False),
            ('H1_10','StackLock 열린 변 합',False),
            ('H1_11','StackLock 개수',False),
            ('H1_12','잠금 해제 기준 합',False),
            ('H1_13','Ads 열린 변 합',True),
            ('H1_14','Ads 수 (최대 3)',True),
            ('H1_15','기믹 열린 변 합',True),
        ]

        if 'h1_weights' not in st.session_state:
            st.session_state.h1_weights = {
                'H1_1':8,'H1_2':12,'H1_3':10,'H1_4':8,'H1_5':10,
                'H1_6':12,'H1_7':12,'H1_8':8,'H1_9':8,'H1_10':5,
                'H1_11':5,'H1_12':6,'H1_13':4,'H1_14':4,'H1_15':4
            }

        total_w = sum(st.session_state.h1_weights.values())
        st.markdown(f"**현재 총합**: {total_w}pt → 각 항목 비율로 환산")

        wc1,wc2 = st.columns(2)
        for i,(key,label,inv) in enumerate(H1_INFO):
            col = wc1 if i%2==0 else wc2
            pct = round(st.session_state.h1_weights[key]/total_w*100,1) if total_w>0 else 0
            new_w = col.slider(
                f"{key} — {label} {'↘역수' if inv else '↗'}",
                0,20,st.session_state.h1_weights[key],key=f"w_{key}"
            )
            st.session_state.h1_weights[key] = new_w
            col.caption(f"{"Ratio" if IS_EN else "비율"}: **{pct}%**")

        # 실시간 파이 차트
        w_vals = list(st.session_state.h1_weights.values())
        w_keys = list(st.session_state.h1_weights.keys())
        fig_pie = go.Figure(go.Pie(
            labels=w_keys, values=w_vals,
            hole=0.4, textinfo='label+percent',
            textfont=dict(size=10)
        ))
        fig_pie.update_layout(height=350,paper_bgcolor=T["plot_bg"],
                              font_color=T["text"],margin=dict(l=10,r=10,t=10,b=10),
                              showlegend=False)
        st.plotly_chart(fig_pie,use_container_width=True)

        # 가중치 설정 JSON 다운로드
        w_config = {"h1_weights": st.session_state.h1_weights,
                    "w_board": st.session_state.w_board,
                    "w_gameplay": st.session_state.w_gameplay}
        st.download_button("📥 가중치 설정 JSON 다운로드",
                           json.dumps(w_config,ensure_ascii=False,indent=2).encode(),
                           "weight_config.json","application/json",use_container_width=True)

    # ── tblStage 가중치 설정
    with set_tab2:
        st.caption(TR("tbl_weight_cap"))

        STACK_INFO = [
            ('alloc',   'TotalAllocation (할당량)',        False, 10, 300),
            ('init_c',  '초기 색상 수',                    False,  1,   5),
            ('dist_c',  '스택당 색상 수',                  False,  1,   4),
            ('dup_r',   '색 중복 확률 (역수 — 낮을수록↑)', True,  0.1, 0.8),
            ('prog1',   '첫 추가 임계값 (역수 — 낮을수록↑)',True,  2,  30),
            ('new_c',   '추가 색상 수',                    False,  0,   5),
            ('gimmick', '기믹 비율',                       False,  0, 0.5),
        ]

        if 'stack_weights' not in st.session_state:
            st.session_state.stack_weights = {
                'alloc':20,'init_c':12,'dist_c':10,
                'dup_r':8,'prog1':10,'new_c':8,'gimmick':14
            }

        tw_s = sum(st.session_state.stack_weights.values())
        st.markdown(f"**{"Total weight" if IS_EN else "Total Weight" if IS_EN else "현재 총합"}**: {tw_s}pt")

        sc1, sc2 = st.columns(2)
        for i, (key, label, inv, mn, mx) in enumerate(STACK_INFO):
            col = sc1 if i%2==0 else sc2
            pct = round(st.session_state.stack_weights[key]/tw_s*100,1) if tw_s>0 else 0
            new_w = col.slider(
                f"{label} {'↘역수' if inv else '↗'}",
                0, 30, st.session_state.stack_weights[key], key=f"sw_{key}"
            )
            st.session_state.stack_weights[key] = new_w
            col.caption(f"{"Ratio" if IS_EN else "비율"}: **{pct}%**")

        # 파이 차트
        sw_vals = list(st.session_state.stack_weights.values())
        sw_keys = ['TotalAlloc','초기색상','스택당색','중복확률','첫임계값','추가색상','기믹비율']
        fig_spie = go.Figure(go.Pie(
            labels=sw_keys, values=sw_vals,
            hole=0.4, textinfo='label+percent',
            textfont=dict(size=10)
        ))
        fig_spie.update_layout(
            height=320, paper_bgcolor=T["plot_bg"],
            font_color=T["text"], margin=dict(l=10,r=10,t=10,b=10),
            showlegend=False
        )
        st.plotly_chart(fig_spie, use_container_width=True)

    # ── 스택 파라미터 수정
    with set_tab3:
        tbl = st.session_state.tbl_df
        if tbl is None:
            st.warning("Please upload tblStage_500.xlsx from the sidebar." if IS_EN else "사이드바에서 tblStage_500.xlsx를 업로드해주세요.")
        else:
            st.caption(TR("stack_cap"))
            show_cols = st.multiselect(
                "표시할 컬럼",
                tbl.columns.tolist(),
                default=[c for c in ['LevelName','TotalAllocation','InitialAvailableColors',
                                     'DistinctColorCount','ColorDuplicationRate',
                                     'ProgressAddNewColor','NewColorsMilestones','DifficultyScore']
                         if c in tbl.columns]
            )
            if show_cols:
                edited = st.data_editor(tbl[show_cols], use_container_width=True, height=500,
                                        num_rows="fixed")
                # 수정된 내용 session_state 반영
                for col in show_cols:
                    st.session_state.tbl_df[col] = edited[col]

                # CSV 다운로드
                st.download_button("📥 CSV Download" if IS_EN else "📥 CSV 다운로드",df_to_csv_bytes(edited),
                                   "tblStage_edited.csv","text/csv",use_container_width=True)

# ══════════════════════════════════════════════════════
# 탭 5 — 아카이브
# ══════════════════════════════════════════════════════
elif page == "🗄️ 6. 아카이브":
    st.markdown('<div class="page-anim">', unsafe_allow_html=True)
    st.title(TR("archive_title"))
    st.caption(TR("archive_cap"))

    ac1, ac2 = st.columns([1,1])

    with ac1:
        st.subheader(TR("save_settings"))
        version_name = st.text_input("Version Name" if IS_EN else TR("version_name"), f"v_{datetime.now().strftime('%Y%m%d_%H%M')}")
        version_memo = st.text_area("Memo" if IS_EN else TR("memo"), placeholder="예: H1-7 강화 테스트, board 가중치 60%로 조정")

        save_payload = {
            "version": version_name,
            "memo": version_memo,
            "timestamp": datetime.now().isoformat(),
            "w_board": st.session_state.w_board,
            "w_gameplay": st.session_state.w_gameplay,
            "h1_weights": st.session_state.get('h1_weights', {}),
        }
        payload_bytes = json.dumps(save_payload, ensure_ascii=False, indent=2).encode("utf-8")

        st.download_button("📥 로컬에 저장 (JSON)", payload_bytes,
                           f"{version_name}.json","application/json",use_container_width=True)

        st.markdown("")
        if st.button("☁️ Commit to GitHub" if IS_EN else TR("commit_github"), use_container_width=True):
            token = st.session_state.github_token
            if not token:
                st.error("사이드바 → 🔑 GitHub 설정에서 토큰을 먼저 입력해주세요.")
            else:
                path = f"{ARCHIVE_PATH}/{version_name}.json"
                with st.spinner("GitHub에 커밋 중..."):
                    ok, resp = github_commit(
                        token, GITHUB_REPO, path, payload_bytes,
                        f"archive: {version_name} — {version_memo[:50]}"
                    )
                if ok:
                    st.success(f"✅ 커밋 완료! `{path}`")
                    st.session_state.archives.append(save_payload)
                else:
                    st.error(f"❌ 커밋 실패: {resp.get('message','')}")

    with ac2:
        st.subheader(TR("load_version"))
        up_archive = st.file_uploader("Upload saved version JSON" if IS_EN else "Upload saved version JSON" if IS_EN else "저장된 버전 JSON 업로드", type=["json"], key="up_arch")
        if up_archive:
            arch = json.load(up_archive)
            st.json(arch)
            if st.button("이 버전으로 복원"):
                st.session_state.w_board    = arch.get('w_board', 50)
                st.session_state.w_gameplay = arch.get('w_gameplay', 50)
                if 'h1_weights' in arch:
                    st.session_state.h1_weights = arch['h1_weights']
                st.success(f"✅ {arch.get('version','')} 복원 완료!")
                st.rerun()

    st.markdown("---")

    # 이번 세션 저장 기록
    if st.session_state.archives:
        st.subheader(TR("session_history"))
        arch_df = pd.DataFrame([{
            '버전': a['version'],
            '저장 시각': a['timestamp'][:19],
            '판 모양 %': a['w_board'],
            '게임 진행 %': a['w_gameplay'],
            '메모': a['memo'],
        } for a in st.session_state.archives])
        st.dataframe(arch_df, use_container_width=True)

        # 버전 비교
        if len(st.session_state.archives) >= 2:
            st.subheader(TR("version_compare"))
            intg = st.session_state.intg_df
            if intg is not None:
                sel_versions = st.multiselect("Select versions to compare (max 3)" if IS_EN else "비교할 버전 선택 (최대 3개)",
                    [a['version'] for a in st.session_state.archives],
                    default=[a['version'] for a in st.session_state.archives[-2:]])
                fig_cmp=go.Figure()
                colors=['#58a6ff','#f78166','#3fb950','#d2a8ff']
                for i,vname in enumerate(sel_versions[:3]):
                    arch_v = next((a for a in st.session_state.archives if a['version']==vname),None)
                    if arch_v:
                        wb=arch_v['w_board']; wg=arch_v['w_gameplay']
                        c=(intg['board_score']*wb+intg['gameplay_score']*wg)/100
                        cs=c.rolling(5,center=True,min_periods=1).mean()
                        fig_cmp.add_trace(go.Scatter(
                            x=list(range(1,len(intg)+1)),y=cs.tolist(),
                            name=f"{vname} (board:{wb}%)",
                            line=dict(color=colors[i],width=2)
                        ))
                fig_cmp.update_layout(height=350,plot_bgcolor=T["plot_bg"],paper_bgcolor=T["plot_bg"],
                    font_color=T["text"],xaxis_title='레벨',
                    xaxis=dict(gridcolor=T["grid_line"]),yaxis=dict(range=[0,105],gridcolor=T["grid_line"]),
                    legend=dict(orientation='h',y=1.1,bgcolor='rgba(0,0,0,0)'),
                    margin=dict(l=10,r=10,t=30,b=10))
                st.plotly_chart(fig_cmp,use_container_width=True)
            else:
                st.info("integrated_difficulty.csv를 사이드바에서 업로드하면 곡선 비교가 가능합니다.")


# ══════════════════════════════════════════════════════
# 탭 7 — 묘수풀이 특수 퍼즐 생성기
# ══════════════════════════════════════════════════════
elif page == "🧩 7. 묘수풀이 생성기":
    import sys as _sys
    _sys.path.insert(0, str(BASE))
    try:
        from generate_special import (
            generate_special_puzzle, analyze_special, solve,
            Board, get_neighbors, COLOR_MAP as SP_COLOR_MAP,
            MATCH_TARGET,
        )
        _HAS_GEN = True
    except ImportError:
        _HAS_GEN = False

    st.markdown('<div class="page-anim">', unsafe_allow_html=True)
    st.markdown('<div class="section-header">🧩 묘수풀이(특수 퍼즐) 생성기</div>', unsafe_allow_html=True)

    if not _HAS_GEN:
        st.error("generate_special.py 파일이 app.py와 같은 디렉토리에 없습니다.")
        st.stop()

    CHIP_COLOR_HEX = {
        0:'#1890FF', 1:'#FADB14', 2:'#F5222D', 3:'#52C41A',
        4:'#FA8C16', 5:'#722ED1', 6:'#BFBFBF', 7:'#141414'
    }
    SP_COLOR_NAMES = ['Blue','Yellow','Red','Green','Orange','Purple','White','Black']

    # ── session state 초기화
    for k, v in {
        'sp_results': [],
        'sp_sim_board': None,
        'sp_sim_hand': [],
        'sp_sim_used': set(),
        'sp_sim_sel': None,
        'sp_sim_history': [],
        'sp_cur_pid': None,
    }.items():
        if k not in st.session_state:
            st.session_state[k] = v

    # ══════════════════════════════════════
    # 섹션 A: 생성 파라미터
    # ══════════════════════════════════════
    st.markdown("#### ① 생성 파라미터")
    col1, col2, col3 = st.columns(3)
    with col1:
        sp_start = st.number_input("시작 번호", min_value=1, max_value=999, value=35)
    with col2:
        sp_end   = st.number_input("끝 번호",   min_value=1, max_value=999, value=37)
    with col3:
        sp_diff  = st.selectbox("난이도", ["D8","D12","D34","D48","D52"], index=2,
                                  help="D8=쉬움(빈칸6) ~ D52=어려움(빈칸2+가교4)")

    sp_seed = st.number_input("시드 (0=자동)", min_value=0, value=0)

    if st.button("🎲 특수 퍼즐 생성", type="primary", use_container_width=True):
        st.session_state.sp_results = []
        progress = st.progress(0)
        status   = st.empty()
        total = sp_end - sp_start + 1
        for i, pid in enumerate(range(int(sp_start), int(sp_end)+1)):
            status.text(f"S_{pid:02d} 생성 중... ({i+1}/{total})")
            try:
                seed = int(sp_seed) if sp_seed > 0 else pid * 12345
                r = generate_special_puzzle(
                    puzzle_id=pid,
                    difficulty=sp_diff,
                    n_colors=3,
                    seed=seed,
                )
                r['pid'] = pid
                r['diff'] = analyze_special(r)
                st.session_state.sp_results.append(r)
            except Exception as e:
                st.session_state.sp_results.append({'pid': pid, 'error': str(e)})
            progress.progress((i+1)/total)
        status.text("완료!")

    # ══════════════════════════════════════
    # 섹션 B: 결과 목록
    # ══════════════════════════════════════
    results = st.session_state.sp_results
    if results:
        st.markdown("---")
        st.markdown("#### ② 생성 결과")

        # 요약 테이블
        rows_disp = []
        for r in results:
            if 'error' in r:
                rows_disp.append({'번호': f"S_{r['pid']:02d}", '상태': '❌ 실패',
                                   '오류': r['error'], '난이도': '-',
                                   '보드칩': '-', '손패칩': '-'})
            else:
                total = r['board_chips'] + r['hand_chips']
                ok = all(v % 10 == 0 for v in total.values())
                rows_disp.append({
                    '번호':    f"S_{r['pid']:02d}",
                    '상태':    '✅ 성공' if ok else '⚠ 불변식오류',
                    '난이도':  f"{r['diff']['diff_score']}점 ({r['diff']['difficulty']})",
'forcing': r['diff']['forcing'],
                    '보드칩':  str(dict(r['board_chips'])),
                    '손패칩':  str(dict(r['hand_chips'])),
                    'Normal':  r.get('normal_cells', '-'),
                    'Stack':   r.get('n_stacks', '-'),
                })
        st.dataframe(pd.DataFrame(rows_disp), use_container_width=True, hide_index=True)

        # 다운로드: JSON ZIP
        import zipfile, io as _io
        zip_buf = _io.BytesIO()
        with zipfile.ZipFile(zip_buf, 'w', zipfile.ZIP_DEFLATED) as zf:
            for r in results:
                if 'error' not in r:
                    zf.writestr(
                        f"S_{r['pid']:03d}.json",
                        json.dumps(r['board_json'], ensure_ascii=False, indent=2)
                    )
        st.download_button(
            "📥 보드 JSON ZIP 다운로드",
            zip_buf.getvalue(),
            f"special_puzzles_S{int(sp_start):02d}-S{int(sp_end):02d}.zip",
            "application/zip",
            use_container_width=True,
        )

        # ── StackInfo / Stage 미리보기
        with st.expander("📋 StackInfo / Stage 행 미리보기"):
            si_rows = [r['stack_info'] for r in results if 'error' not in r]
            st_rows = [r['stage_row']  for r in results if 'error' not in r]
            if si_rows:
                st.markdown("**StackInfo 탭**")
                st.dataframe(pd.DataFrame(si_rows), use_container_width=True, hide_index=True)
                st.markdown("**Stage 탭**")
                st.dataframe(pd.DataFrame(st_rows), use_container_width=True, hide_index=True)

        # ── tblStage xlsx 업데이트
        st.markdown("---")
        st.markdown("#### ③ tblStage.xlsx 업데이트")
        up_tbl_sp = st.file_uploader("tblStage.xlsx 업로드 (업데이트할 파일)", type=["xlsx"], key="up_tbl_sp")
        if up_tbl_sp and st.button("📝 xlsx에 행 추가 및 다운로드", use_container_width=True):
            from openpyxl import load_workbook as _lwb
            wb2 = _lwb(up_tbl_sp)
            ws_s  = wb2['Stage']
            ws_si = wb2['StackInfo']
            s_headers  = [c.value for c in next(ws_s.iter_rows(min_row=1, max_row=1))]
            si_headers = [c.value for c in next(ws_si.iter_rows(min_row=1, max_row=1))]

            exist_ids = {row[0] for row in ws_s.iter_rows(min_row=2, values_only=True) if row[0]}
            exist_si  = {row[0] for row in ws_si.iter_rows(min_row=2, values_only=True) if row[0]}

            added = 0
            for r in results:
                if 'error' in r: continue
                pid = r['pid']
                if (1000+pid) in exist_ids or pid in exist_si: continue

                row_num = ws_s.max_row + 1
                sr = r['stage_row']
                stage_vals = []
                for h in s_headers:
                    if h == 'Id':                        stage_vals.append(1000+pid)
                    elif h == 'Mode':                    stage_vals.append('Turn')
                    elif h == 'LevelName':               stage_vals.append(f'S {pid:02d}')
                    elif h == 'PlaceableCount':          stage_vals.append(3)
                    elif h == 'IsPreview':               stage_vals.append(False)
                    elif h == 'TotalAllocation':         stage_vals.append(sr['TurnCount'])
                    elif h == 'Extra':                   stage_vals.append(pid)
                    elif h == 'TurnCount':               stage_vals.append(sr['TurnCount'])
                    elif h in ('IceCount','GrassCount','WoodCount','CameraPictureCount'): stage_vals.append(0)
                    elif h == 'GenreXPReward':           stage_vals.append(10)
                    elif h == 'XpReward':    stage_vals.append(f'=CEILING(65+((A{row_num}-1000)^ 1.3), 5)')
                    elif h == 'GoldReward':  stage_vals.append(f'=CEILING(800+((A{row_num}-1002)^ 3), 5)')
                    elif h == 'TokenReward': stage_vals.append(f'=CEILING(10+((A{row_num}-1000)^ 1.5), 5)')
                    elif h == 'GemReward':   stage_vals.append(f'=CEILING(5+((A{row_num}-1000)^ 1.45), 5)')
                    else:                                stage_vals.append(None)
                ws_s.append(stage_vals)

                si_vals = []
                si_data = r['stack_info']
                for h in si_headers:
                    if h == 'Id': si_vals.append(pid)
                    elif h and h.startswith('Stack'): si_vals.append(si_data.get(h))
                    else: si_vals.append(None)
                ws_si.append(si_vals)
                added += 1

            buf2 = _io.BytesIO()
            wb2.save(buf2)
            st.success(f"{added}개 행 추가 완료!")
            st.download_button(
                "📥 tblStage_updated.xlsx 다운로드",
                buf2.getvalue(),
                "tblStage_updated.xlsx",
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                use_container_width=True,
            )

    # ══════════════════════════════════════
    # 섹션 C: 인게임 시뮬레이터
    # ══════════════════════════════════════
    st.markdown("---")
    st.markdown("#### ④ 퍼즐 시뮬레이터")

    if not results or all('error' in r for r in results):
        st.info("위에서 퍼즐을 먼저 생성하세요.")
    else:
        ok_results = [r for r in results if 'error' not in r]
        sel_label = st.selectbox(
            "시뮬레이션할 퍼즐 선택",
            [f"S_{r['pid']:02d}  ({r['diff']['difficulty']} {r['diff']['diff_score']}점)" for r in ok_results],
            key="sp_sel_puzzle",
        )
        sel_idx = [f"S_{r['pid']:02d}  ({r['diff']['difficulty']} {r['diff']['diff_score']}점)" for r in ok_results].index(sel_label)
        sel_r = ok_results[sel_idx]

        if st.button("🔄 이 퍼즐 불러오기", key="sp_load"):
            tiles = sel_r['board_json']['Tiles']
            Y = sel_r['board_json']['YCells']
            X = sel_r['board_json']['XCells']
            grid = []
            for y in range(Y):
                row = []
                for x in range(X):
                    tt = tiles[y][x].get('TileType', 1)
                    if   tt == 1: row.append(None)
                    elif tt == 0: row.append([])
                    else:          row.append(list(tiles[y][x].get('Stacks', [])))
                grid.append(row)
            st.session_state.sp_sim_board   = grid
            st.session_state.sp_sim_hand    = [list(s) for s in sel_r['hand_stacks']]
            st.session_state.sp_sim_used    = set()
            st.session_state.sp_sim_sel     = None
            st.session_state.sp_sim_history = []
            st.session_state.sp_cur_pid     = sel_r['pid']

        board_grid = st.session_state.sp_sim_board
        if board_grid is not None:
            Y_b = len(board_grid); X_b = len(board_grid[0])
            hand_stacks_sim = st.session_state.sp_sim_hand
            used = st.session_state.sp_sim_used
            sel_hand = st.session_state.sp_sim_sel

            # ── 보드 시각화 (Plotly)
            import plotly.graph_objects as _go
            import math as _math

            HEX_R = 35
            fig = _go.Figure()
            fig.update_layout(
                width=500, height=350,
                margin=dict(l=10,r=10,t=10,b=10),
                xaxis=dict(visible=False, range=[-20, X_b*HEX_R*1.6+20]),
                yaxis=dict(visible=False, scaleanchor='x', range=[-20, Y_b*HEX_R*1.6+20]),
                plot_bgcolor=T['plot_bg'], paper_bgcolor=T['plot_bg'],
                showlegend=False,
            )

            for y in range(Y_b):
                for x in range(X_b):
                    cell = board_grid[y][x]
                    if cell is None: continue
                    cx = x * HEX_R*1.5 + (HEX_R*0.75 if y%2 else 0) + HEX_R
                    cy = y * HEX_R*0.87*2 + HEX_R

                    pts_x = [cx + HEX_R*_math.cos(_math.radians(60*i-30)) for i in range(6)]
                    pts_y = [cy + HEX_R*_math.sin(_math.radians(60*i-30)) for i in range(6)]
                    pts_x.append(pts_x[0]); pts_y.append(pts_y[0])

                    is_empty = len(cell) == 0
                    fill_col = '#D5E8F5' if (is_empty and sel_hand is not None) else (
                               T['bg2'] if is_empty else T['bg3'])
                    border_col = T['border'] if not is_empty else (
                                 T['brown'] if sel_hand is not None else T['brown_lt'])

                    fig.add_trace(_go.Scatter(
                        x=pts_x, y=pts_y, fill='toself',
                        fillcolor=fill_col,
                        line=dict(color=border_col, width=2 if (is_empty and sel_hand is not None) else 1),
                        mode='lines', hoverinfo='skip',
                    ))

                    # 칩 표시 (top부터 최대 3개)
                    if cell:
                        show = cell[-3:]  # [-3:]= 위 3개, [-1]=top
                        for ki, c_code in enumerate(reversed(show)):  # top이 위에 오도록
                            dot_y = cy + 8 - ki*12
                            fig.add_trace(_go.Scatter(
                                x=[cx], y=[dot_y],
                                mode='markers+text',
                                marker=dict(size=14, color=CHIP_HEX.get(c_code,'#888'),
                                            line=dict(color='white', width=1)),
                                text=[SP_COLOR_NAMES[c_code][0]] if ki==0 else [''],
                                textposition='middle center',
                                textfont=dict(size=8, color='white'),
                                hoverinfo='skip',
                            ))
                        # 총 개수
                        fig.add_trace(_go.Scatter(
                            x=[cx+22], y=[cy+22],
                            mode='text', text=[str(len(cell))],
                            textfont=dict(size=9, color=T['text2']),
                            hoverinfo='skip',
                        ))
                    else:
                        # 빈 칸 표시
                        fig.add_trace(_go.Scatter(
                            x=[cx], y=[cy],
                            mode='text', text=['○'],
                            textfont=dict(size=20, color=T['brown'] if sel_hand is not None else T['brown_lt']),
                            hoverinfo='skip',
                        ))

            st.plotly_chart(fig, use_container_width=False)

            # ── 칸 선택 (배치용)
            if sel_hand is not None:
                st.markdown(f"**Stack {sel_hand+1}** 선택됨 — 배치할 빈 칸 좌표 입력:")
                c1, c2, c3 = st.columns([1,1,2])
                with c1:
                    place_y = st.number_input("행(Y)", 0, Y_b-1, 0, key="sp_py")
                with c2:
                    place_x = st.number_input("열(X)", 0, X_b-1, 0, key="sp_px")
                with c3:
                    st.write("")
                    if st.button("✅ 배치", key="sp_place"):
                        py, px = int(place_y), int(place_x)
                        cell = board_grid[py][px]
                        if cell is None:
                            st.warning("Blank 칸입니다.")
                        elif len(cell) > 0:
                            st.warning("이미 칩이 있는 칸입니다.")
                        else:
                            # 히스토리 저장
                            import copy
                            st.session_state.sp_sim_history.append({
                                'board': copy.deepcopy(board_grid),
                                'used':  set(used),
                                'sel':   sel_hand,
                            })
                            # Board 객체로 배치+cascade
                            b = Board(board_grid)
                            b.place(py, px, hand_stacks_sim[sel_hand])
                            st.session_state.sp_sim_board = b.g
                            st.session_state.sp_sim_used.add(sel_hand)
                            st.session_state.sp_sim_sel = None

                            # 클리어/실패 판정
                            if len(st.session_state.sp_sim_used) == len(hand_stacks_sim):
                                if b.all_clear():
                                    st.success("🎉 클리어! 보드가 완전히 비워졌어요!")
                                else:
                                    st.error("손패를 모두 사용했지만 보드에 칩이 남아있어요.")
                            st.rerun()

            # ── 손패 표시
            st.markdown("**손패**")
            hand_cols = st.columns(len(hand_stacks_sim))
            for hi, hs in enumerate(hand_stacks_sim):
                with hand_cols[hi]:
                    is_used = hi in used
                    is_sel  = sel_hand == hi
                    border  = "3px solid #6B3A2A" if is_sel else "1px solid #C4956A"
                    opacity = "0.35" if is_used else "1.0"
                    chip_html = ''.join(
                        f'<span style="display:inline-block;width:16px;height:16px;border-radius:50%;'
                        f'background:{CHIP_HEX.get(c,"#888")};margin:1px;"></span>'
                        for c in reversed(hs)  # top이 먼저 보이도록
                    )
                    st.markdown(
                        f'<div style="border:{border};border-radius:8px;padding:8px;'
                        f'text-align:center;opacity:{opacity};background:#fff;">'
                        f'<div style="font-size:11px;color:#7A5C45;">Stack {hi+1}</div>'
                        f'<div style="margin:4px 0;">{chip_html}</div>'
                        f'<div style="font-size:10px;color:#999;">{len(hs)}칩</div>'
                        f'</div>',
                        unsafe_allow_html=True,
                    )
                    if not is_used:
                        if st.button(f"선택" if not is_sel else "선택 해제",
                                     key=f"sp_hand_{hi}"):
                            st.session_state.sp_sim_sel = hi if not is_sel else None
                            st.rerun()

            # ── 되돌리기
            col_u, col_sol = st.columns(2)
            with col_u:
                if st.button("↩ 되돌리기", key="sp_undo"):
                    if st.session_state.sp_sim_history:
                        snap = st.session_state.sp_sim_history.pop()
                        st.session_state.sp_sim_board = snap['board']
                        st.session_state.sp_sim_used  = snap['used']
                        st.session_state.sp_sim_sel   = snap['sel']
                        st.rerun()
                    else:
                        st.info("되돌릴 동작이 없어요.")
            with col_sol:
                if st.button("✨ 정답 보기", key="sp_auto"):
                    tiles = sel_r['board_json']['Tiles']
                    sol = sel_r.get('solution')
                    if sol:
                        steps = []
                        for step in sol:
                            hi = step['hand_idx']
                            py, px = step['pos']
                            chips = [SP_COLOR_NAMES[c] for c in step['chips']]
                            steps.append(f"Stack {hi+1} [{','.join(chips)}] → ({py},{px})")
                        st.info("정답 배치 순서:\n" + "\n".join(f"{i+1}. {s}" for i,s in enumerate(steps)))
                    else:
                        st.warning("정답 정보가 없습니다.")

    st.markdown('</div>', unsafe_allow_html=True)
