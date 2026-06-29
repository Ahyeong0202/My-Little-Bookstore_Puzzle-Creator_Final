"""
update_tblstage.py (v2)
──────────────────
generate_special.py(v4)로 생성한 묘수풀이 퍼즐들을 tblStage.xlsx에 추가.

사용법:
  python3 update_tblstage.py <tblStage.xlsx 경로> <시작번호> <끝번호> [출력경로]

예:
  python3 update_tblstage.py data/tblStage.xlsx 35 40
  → data/tblStage_updated.xlsx 생성

추가 내용:
  - Stage 탭: Mode=Turn 행 추가 (Id=1000+pid)
  - StackInfo 탭: Stack1~3 행 추가 (Id=pid)

이번 버전(v2)에서 바뀐 점:
  - 난이도는 더 이상 _get_difficulty가 아니라 generate_special.difficulty_for_id
    (15개 순환 패턴: [2,4,2,3,2,3,2,3,1,4,1,4,3,5,2])로 자동 결정된다.
  - generate_special_puzzle은 이제 n_colors 인자를 받지 않는다
    (난이도에 따라 자동으로 3~4색을 시도한다).
  - r['board_chips']/r['hand_chips']/r['stage_row']/r['diff_score'] 필드가
    없어졌으므로, 이 스크립트 안에서 직접 계산/구성한다.
  - board_json도 generate_special.py(v4)의 새 좌표계/포맷(zip 샘플과 동일,
    GameType:0)을 그대로 사용한다.
"""

import sys
import json
from pathlib import Path
from collections import Counter
from openpyxl import load_workbook

sys.path.insert(0, str(Path(__file__).parent))
from generate_special import generate_special_puzzle, analyze_special, difficulty_for_id


# ── 보상 공식 생성 (기존 패턴 유지)
def _xp_formula(row_num):
    return f'=CEILING(65+((A{row_num}-1000)^ 1.3), 5)'


def _gold_formula(row_num):
    return f'=CEILING(800+((A{row_num}-1002)^ 3), 5)'


def _token_formula(row_num):
    return f'=CEILING(10+((A{row_num}-1000)^ 1.5), 5)'


def _gem_formula(row_num):
    return f'=CEILING(5+((A{row_num}-1000)^ 1.45), 5)'


def _hand_chip_counts(hand_stacks):
    c = Counter()
    for h in hand_stacks:
        for chip in h:
            c[chip] += 1
    return c


def update_tblstage(
    xlsx_path: str,
    start: int,
    end: int,
    output_path: str = None,
    seed_base: int = 12345,
):
    """
    tblStage.xlsx에 특수 퍼즐 행 추가.

    Parameters
    ----------
    xlsx_path   : 원본 tblStage.xlsx 경로
    start, end  : 추가할 S 번호 범위 (start~end 포함)
    output_path : 저장 경로 (None이면 _updated 접미사)
    """
    src = Path(xlsx_path)
    if output_path is None:
        output_path = str(src.parent / (src.stem + '_updated.xlsx'))

    wb = load_workbook(str(src))
    ws_stage = wb['Stage']
    ws_stack = wb['StackInfo']

    # ── Stage 탭 헤더 파악
    stage_headers = [c.value for c in next(ws_stage.iter_rows(min_row=1, max_row=1))]
    stage_last_row = ws_stage.max_row

    # StackInfo 헤더
    stack_headers = [c.value for c in next(ws_stack.iter_rows(min_row=1, max_row=1))]

    # ── 기존 Id 확인 (중복 방지)
    existing_stage_ids = set()
    for row in ws_stage.iter_rows(min_row=2, values_only=True):
        if row[0]:
            existing_stage_ids.add(row[0])

    existing_stack_ids = set()
    for row in ws_stack.iter_rows(min_row=2, values_only=True):
        if row[0]:
            existing_stack_ids.add(row[0])

    print(f'기존 Stage 행: {len(existing_stage_ids)}개 '
          f'(Turn: {sum(1 for i in existing_stage_ids if isinstance(i, int) and i >= 1000)}개)')
    print(f'기존 StackInfo 행: {len(existing_stack_ids)}개')
    print()

    added_stage = 0
    added_stack = 0
    generated = []

    for pid in range(start, end + 1):
        stage_id = 1000 + pid
        stack_id = pid

        if stage_id in existing_stage_ids:
            print(f'  S_{pid:02d}: Stage Id={stage_id} 이미 존재 — 스킵')
            continue
        if stack_id in existing_stack_ids:
            print(f'  S_{pid:02d}: StackInfo Id={stack_id} 이미 존재 — 스킵')
            continue

        diff_level = difficulty_for_id(pid)

        try:
            r = generate_special_puzzle(
                puzzle_id=pid,
                difficulty=diff_level,
                seed=pid * seed_base,
            )
        except RuntimeError as e:
            print(f'  S_{pid:02d}: 생성 실패 — {e}')
            continue

        diff = analyze_special(r)
        hand_counts = _hand_chip_counts(r['hand_stacks'])

        # ── Stage 탭에 행 추가
        stage_row_num = stage_last_row + added_stage + 1

        row_vals = []
        for h in stage_headers:
            if h == 'Id':
                row_vals.append(stage_id)
            elif h == 'Mode':
                row_vals.append('Turn')
            elif h == 'LevelName':
                row_vals.append(f'S {pid:02d}')
            elif h == 'PlaceableCount':
                row_vals.append(3)
            elif h == 'IsPreview':
                row_vals.append(False)
            elif h == 'TotalAllocation':
                row_vals.append(3)  # 특수 퍼즐 고정 (손패 3장)
            elif h == 'InitialAvailableColors':
                row_vals.append(None)
            elif h == 'DistinctColorCount':
                row_vals.append(None)
            elif h == 'ColorDuplicationRate':
                row_vals.append(None)
            elif h == 'ProgressAddNewColor':
                row_vals.append(None)
            elif h == 'NewColorsMilestones':
                row_vals.append(None)
            elif h == 'Extra':
                row_vals.append(pid)  # StackInfo 세트 번호
            elif h == 'TurnCount':
                row_vals.append(3)
            elif h == 'IceCount':
                row_vals.append(0)
            elif h == 'GrassCount':
                row_vals.append(0)
            elif h == 'WoodCount':
                row_vals.append(0)
            elif h == 'CameraPictureCount':
                row_vals.append(0)
            elif h == 'GenreXPReward':
                row_vals.append(10)
            elif h == 'XpReward':
                row_vals.append(_xp_formula(stage_row_num))
            elif h == 'GoldReward':
                row_vals.append(_gold_formula(stage_row_num))
            elif h == 'TokenReward':
                row_vals.append(_token_formula(stage_row_num))
            elif h == 'GemReward':
                row_vals.append(_gem_formula(stage_row_num))
            else:
                row_vals.append(None)

        ws_stage.append(row_vals)
        added_stage += 1

        # ── StackInfo 탭에 행 추가
        si = r['stack_info']
        stack_vals = []
        for h in stack_headers:
            if h == 'Id':
                stack_vals.append(stack_id)
            elif h and h.startswith('Stack'):
                stack_vals.append(si.get(h))
            else:
                stack_vals.append(None)
        ws_stack.append(stack_vals)
        added_stack += 1

        total_ok = all(v == 10 for v in r['total_counts'].values())
        print(f"  S_{pid:02d} [난이도{r['difficulty']} 색{r['n_colors']}] "
              f"forcing={r['forcing']} sol={r['n_solutions']} "
              f"보드={r['board_counts']} 손패={dict(hand_counts)} "
              f"{'✓' if total_ok else '✗'}")

        generated.append({
            'pid': pid,
            'board_json': r['board_json'],
            'stack_info': r['stack_info'],
            'difficulty': r['difficulty'],
            'diff': diff,
        })

    # ── 저장
    wb.save(output_path)
    print()
    print(f'저장 완료: {output_path}')
    print(f'  Stage 추가: {added_stage}행')
    print(f'  StackInfo 추가: {added_stack}행')
    return generated


def export_json_files(generated, out_dir: str):
    """generated 리스트의 board_json들을 zip 샘플과 동일한 형식(S NN.json)으로 저장."""
    outp = Path(out_dir)
    outp.mkdir(parents=True, exist_ok=True)
    for g in generated:
        fname = outp / f"S {g['pid']:02d}.json"
        with open(fname, 'w', encoding='utf-8') as f:
            json.dump(g['board_json'], f, ensure_ascii=False, indent=2)
    print(f'JSON {len(generated)}개 저장 완료: {outp}')


if __name__ == '__main__':
    if len(sys.argv) < 4:
        print('사용법: python3 update_tblstage.py <tblStage.xlsx> <시작번호> <끝번호> [출력경로]')
        sys.exit(1)

    xlsx = sys.argv[1]
    start = int(sys.argv[2])
    end = int(sys.argv[3])
    out = sys.argv[4] if len(sys.argv) > 4 else None

    result = update_tblstage(xlsx, start, end, out)
    print(f'\n생성된 퍼즐: {len(result)}개')

    # JSON 파일도 함께 저장 (게임에 바로 적용 가능한 형식)
    if result:
        src = Path(xlsx)
        json_dir = src.parent / 'special_json_output'
        export_json_files(result, str(json_dir))
